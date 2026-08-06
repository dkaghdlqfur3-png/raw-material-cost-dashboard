import base64
import html
import io
import json
import re
from datetime import datetime
from urllib.error import HTTPError, URLError
from urllib.parse import quote
from urllib.request import Request, urlopen

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import pytz
import streamlit as st


SAVED_FILE_NAME = "latest.xlsx"
DEFAULT_GITHUB_REPOSITORY = "dkaghdlqfur3-png/raw-material-cost-dashboard"
DEFAULT_GITHUB_BRANCH = "main"


def get_storage_settings():
    """GitHub 저장 설정을 반환하되 토큰은 앱 소스에 노출하지 않습니다."""
    try:
        storage = st.secrets["github_storage"]
        token = str(storage["token"])
        repository = str(storage.get("repository", DEFAULT_GITHUB_REPOSITORY))
        branch = str(storage.get("branch", DEFAULT_GITHUB_BRANCH))
        file_path = str(storage.get("file_path", SAVED_FILE_NAME))
        return token, repository, branch, file_path
    except (KeyError, FileNotFoundError):
        return None


def github_contents_request(method="GET", body=None):
    settings = get_storage_settings()
    if settings is None:
        if method != "GET":
            raise RuntimeError("GitHub 저장 설정이 아직 완료되지 않았습니다.")
        token = None
        repository = DEFAULT_GITHUB_REPOSITORY
        branch = DEFAULT_GITHUB_BRANCH
        file_path = SAVED_FILE_NAME
    else:
        token, repository, branch, file_path = settings
    endpoint = (
        f"https://api.github.com/repos/{quote(repository, safe='/')}/contents/"
        f"{quote(file_path, safe='/')}"
    )
    if method == "GET":
        endpoint += f"?ref={quote(branch, safe='')}"
    headers = {
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "raw-material-cost-dashboard",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    data = None
    if body is not None:
        headers["Content-Type"] = "application/json"
        data = json.dumps(body).encode("utf-8")

    request = Request(endpoint, data=data, headers=headers, method=method)
    try:
        with urlopen(request, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        if method == "GET" and exc.code == 404:
            return None
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"GitHub 응답 오류({exc.code}): {detail}") from exc
    except URLError as exc:
        raise RuntimeError(f"GitHub에 연결하지 못했습니다: {exc.reason}") from exc


def load_saved_file():
    file_info = github_contents_request("GET")
    if file_info is None:
        return None
    encoded_content = str(file_info.get("content", "")).replace("\n", "")
    if not encoded_content:
        raise RuntimeError("GitHub에서 저장 파일의 내용을 읽지 못했습니다.")
    return base64.b64decode(encoded_content)


def load_saved_timestamp():
    """latest.xlsx를 마지막으로 저장한 GitHub 커밋 시각을 한국 시간으로 반환합니다."""
    settings = get_storage_settings()
    if settings is None:
        token = None
        repository = DEFAULT_GITHUB_REPOSITORY
        branch = DEFAULT_GITHUB_BRANCH
        file_path = SAVED_FILE_NAME
    else:
        token, repository, branch, file_path = settings
    endpoint = (
        f"https://api.github.com/repos/{quote(repository, safe='/')}/commits"
        f"?path={quote(file_path, safe='/')}&sha={quote(branch, safe='')}&per_page=1"
    )
    headers = {
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "raw-material-cost-dashboard",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    request = Request(endpoint, headers=headers, method="GET")
    try:
        with urlopen(request, timeout=30) as response:
            commits = json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"GitHub 저장 일시 조회 오류({exc.code}): {detail}") from exc
    except URLError as exc:
        raise RuntimeError(f"GitHub에 연결하지 못했습니다: {exc.reason}") from exc

    if not commits:
        return None
    saved_at_text = commits[0].get("commit", {}).get("committer", {}).get("date")
    if not saved_at_text:
        return None
    saved_at = datetime.fromisoformat(saved_at_text.replace("Z", "+00:00"))
    return saved_at.astimezone(pytz.timezone("Asia/Seoul"))


def save_latest_file(file_bytes):
    settings = get_storage_settings()
    if settings is None:
        raise RuntimeError("GitHub 저장 설정이 아직 완료되지 않았습니다.")
    _, _, branch, _ = settings
    current_file = github_contents_request("GET")
    body = {
        "message": "대시보드 최신 데이터 저장",
        "content": base64.b64encode(file_bytes).decode("ascii"),
        "branch": branch,
    }
    if current_file is not None and current_file.get("sha"):
        body["sha"] = current_file["sha"]
    github_contents_request("PUT", body)


st.set_page_config(
    page_title="원가 영향 인자 대시보드",
    page_icon="📊",
    layout="wide",
)

# =========================================================
# 화면 스타일
# =========================================================
st.markdown(
    """
    <style>
    .block-container {
        max-width: 120rem;
        padding-top: 2.5rem;
        padding-bottom: 2rem;
    }

    h1 {
        font-size: 2.15rem !important;
        line-height: 1.35 !important;
        padding-top: 0.25rem !important;
        margin-top: 0 !important;
        overflow: visible !important;
    }
    h2 { font-size: 1.65rem !important; }
    h3 { font-size: 1.30rem !important; }

    div[data-testid="stMetric"] {
        background-color: #f8f9fb;
        border: 1px solid #e7eaf0;
        padding: 15px;
        border-radius: 14px;
    }

    div[data-testid="stMetricLabel"] {
        font-size: 1.02rem !important;
        font-weight: 700 !important;
    }

    div[data-testid="stMetricValue"] {
        font-size: 1.65rem !important;
        font-weight: 700 !important;
    }

    div[data-testid="stMetricDelta"] {
        font-size: 0.95rem !important;
        font-weight: 600 !important;
    }

    [data-testid="stDataFrame"] td { font-size: 1rem !important; }
    [data-testid="stDataFrame"] th { font-size: 1.03rem !important; }

    .insight-summary {
        border-radius: 16px;
        padding: 20px 22px;
        margin: 8px 0 18px 0;
        border-left: 7px solid;
    }
    .insight-summary.red { background: #fff1f1; border-color: #d32f2f; }
    .insight-summary.blue { background: #eef6ff; border-color: #1976d2; }
    .insight-summary.orange { background: #fff7e8; border-color: #ef6c00; }
    .insight-summary.green { background: #edf8f1; border-color: #2e7d32; }
    .insight-summary-title { font-size: 1.25rem; font-weight: 800; margin-bottom: 7px; }
    .insight-summary-text { font-size: 1rem; line-height: 1.65; color: #344054; }

    .insight-kpi {
        border-radius: 14px;
        padding: 16px 18px;
        min-height: 108px;
        border: 1px solid;
    }
    .insight-kpi.red { background: #fff3f3; border-color: #ffc9c9; color: #c62828; }
    .insight-kpi.blue { background: #eef7ff; border-color: #c5e2ff; color: #1565c0; }
    .insight-kpi.orange { background: #fff7e8; border-color: #ffe0a3; color: #e65100; }
    .insight-kpi.gray { background: #f5f6f8; border-color: #dde1e7; color: #475467; }
    .insight-kpi-label { font-size: .96rem; font-weight: 700; }
    .insight-kpi-value { font-size: 1.8rem; font-weight: 800; margin-top: 9px; }

    .insight-panel {
        border-radius: 14px;
        padding: 18px;
        min-height: 270px;
        border-top: 5px solid;
        background: #ffffff;
        box-shadow: 0 2px 10px rgba(16, 24, 40, .06);
    }
    .insight-panel.red { border-color: #d32f2f; }
    .insight-panel.blue { border-color: #1976d2; }
    .insight-panel.gray { border-color: #7a8391; }
    .insight-panel-title { font-size: 1.15rem; font-weight: 800; margin-bottom: 12px; }
    .insight-row { padding: 10px 0; border-bottom: 1px solid #edf0f4; }
    .insight-row:last-child { border-bottom: none; }
    .insight-row-top { display: flex; justify-content: space-between; gap: 12px; align-items: center; }
    .insight-item-name { font-weight: 800; color: #172033; }
    .insight-change { font-weight: 800; white-space: nowrap; }
    .insight-change.red { color: #d32f2f; }
    .insight-change.blue { color: #1565c0; }
    .insight-change.gray { color: #667085; }
    .insight-related { color: #7a8391; font-size: .86rem; margin-top: 3px; }

    .trend-report {
        background: linear-gradient(135deg, #f8fafc 0%, #eef4fb 100%);
        border: 1px solid #d9e2ef;
        border-left: 7px solid #355c8a;
        border-radius: 16px;
        padding: 22px 24px;
        margin: 8px 0 20px 0;
    }
    .trend-report-title { font-size: 1.25rem; font-weight: 800; color: #172b4d; margin-bottom: 10px; }
    .trend-report-body { font-size: 1rem; line-height: 1.8; color: #344054; }
    .trend-group-card {
        border-radius: 13px;
        padding: 15px 17px;
        min-height: 130px;
        border: 1px solid;
    }
    .trend-group-card.red { background: #fff3f3; border-color: #ffc9c9; }
    .trend-group-card.blue { background: #eef7ff; border-color: #c5e2ff; }
    .trend-group-card.orange { background: #fff7e8; border-color: #ffe0a3; }
    .trend-group-card.gray { background: #f5f6f8; border-color: #dde1e7; }
    .trend-group-title { font-weight: 800; font-size: 1.03rem; margin-bottom: 8px; }
    .trend-group-items { font-weight: 700; color: #25324b; line-height: 1.55; }
    .trend-group-note { color: #667085; font-size: .86rem; margin-top: 6px; }

    .briefing-box {
        background: #172b4d;
        color: white;
        border-radius: 16px;
        padding: 28px 32px;
        margin: 10px 0 24px 0;
    }
    .briefing-box-title { font-size: 1.48rem; font-weight: 850; margin-bottom: 13px; }
    .briefing-line { font-size:1.03rem; line-height: 1.85; color: #edf3fb; }
    .briefing-grid {
        display: grid;
        grid-template-columns: repeat(4, minmax(0, 1fr));
        gap: 14px;
        margin: 8px 0 24px 0;
    }
    .briefing-card {
        background: white;
        border: 1px solid #e2e7ef;
        border-top: 5px solid #98a2b3;
        border-radius: 14px;
        padding: 17px 18px;
        box-shadow: 0 2px 9px rgba(16, 24, 40, .05);
    }
    .briefing-card.up { border-top-color: #d32f2f; }
    .briefing-card.down { border-top-color: #1976d2; }
    .briefing-card.flat { border-top-color: #7a8391; }
    .briefing-card-name { font-size: 1.08rem; font-weight: 800; color: #172033; }
    .briefing-card-related { font-size: .83rem; color: #7a8391; margin: 3px 0 13px 0; }
    .briefing-card-value { display:block; font-size: 1.58rem; line-height:1.25; font-weight: 850; color: #172033; margin-bottom:7px; }
    .briefing-card-change { display: inline-block; font-size: .92rem; font-weight: 800; margin-left: 0; }
    .briefing-card-change.up { color: #d32f2f; }
    .briefing-card-change.down { color: #1976d2; }
    .briefing-card-change.flat { color: #667085; }
    .briefing-card-trend { font-weight: 750; margin-top: 12px; color: #344054; }
    .briefing-card-note { font-size: .88rem; color: #667085; line-height: 1.5; margin-top: 4px; }
    @media (max-width: 1200px) {
        .briefing-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
    }
    @media (max-width: 700px) {
        .briefing-grid { grid-template-columns: 1fr; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# =========================================================
# 기본값
# =========================================================
KST = pytz.timezone("Asia/Seoul")
NOW = datetime.now(KST)

SHEET_ALIASES = {
    "원자재": ["원자재", "raw", "raw material", "raw_material", "commodity", "commodities"],
    "환율": ["환율", "fx", "exchange", "exchange rate", "currency"],
}

COLUMN_ALIASES = {
    "item": ["품목", "항목", "원자재", "통화", "item", "name", "indicator", "지표"],
    "date": ["날짜", "일자", "기준일", "년월", "월", "date", "month", "ym"],
    "value": ["값", "가격", "시세", "환율", "지수", "단가", "value", "price", "rate"],
    "unit": ["단위", "unit"],
    "currency": ["통화", "currency"],
    "related": ["사용 품목", "관련 품목", "적용 품목", "연관 품목", "related", "related item"],
    "source": ["출처", "source"],
    "note": ["비고", "메모", "참고", "note", "memo"],
    "category": ["구분", "분류", "category", "type"],
}


# =========================================================
# 공통 유틸리티
# =========================================================
def normalize_text(value):
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def find_column(df, aliases):
    normalized = {normalize_text(c): c for c in df.columns}
    for alias in aliases:
        key = normalize_text(alias)
        if key in normalized:
            return normalized[key]
    return None


def clean_number(value):
    if pd.isna(value):
        return pd.NA
    if isinstance(value, (int, float, np.number)):
        return float(value)

    text = str(value).strip()
    if not text:
        return pd.NA

    text = text.replace(",", "")
    text = re.sub(r"[₩$€¥%]", "", text)
    text = re.sub(r"\(([^)]+)\)", r"-\1", text)
    text = re.sub(r"[^0-9.\-]", "", text)

    if text in {"", "-", ".", "-."}:
        return pd.NA

    try:
        return float(text)
    except ValueError:
        return pd.NA


def parse_month(value):
    if pd.isna(value):
        return pd.NaT

    if isinstance(value, pd.Timestamp):
        return value.to_period("M").to_timestamp()

    if isinstance(value, datetime):
        return pd.Timestamp(value).to_period("M").to_timestamp()

    # Excel 날짜 일련번호 대응
    if isinstance(value, (int, float, np.number)) and 20000 < float(value) < 80000:
        try:
            dt = pd.Timestamp("1899-12-30") + pd.to_timedelta(float(value), unit="D")
            return dt.to_period("M").to_timestamp()
        except Exception:
            pass

    text = str(value).strip()
    if not text:
        return pd.NaT

    text = text.replace("년", "-").replace("월", "")
    text = text.replace("/", "-").replace(".", "-")
    text = re.sub(r"\s+", "", text)

    patterns = [
        (r"^\d{4}-\d{1,2}$", lambda x: x + "-01"),
        (r"^\d{2}-\d{1,2}$", lambda x: "20" + x + "-01"),
        (r"^\d{6}$", lambda x: x[:4] + "-" + x[4:] + "-01"),
        (r"^\d{4}$", lambda x: x + "-01-01"),
    ]

    for pattern, converter in patterns:
        if re.fullmatch(pattern, text):
            text = converter(text)
            break

    parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        return pd.NaT
    return parsed.to_period("M").to_timestamp()


def looks_like_month_column(column_name):
    return not pd.isna(parse_month(column_name))


def select_sheet(sheet_names, target):
    aliases = [normalize_text(x) for x in SHEET_ALIASES[target]]
    normalized_names = {normalize_text(name): name for name in sheet_names}

    for alias in aliases:
        if alias in normalized_names:
            return normalized_names[alias]

    for name in sheet_names:
        normalized_name = normalize_text(name)
        if any(alias in normalized_name for alias in aliases):
            return name

    return None


# =========================================================
# 엑셀 데이터 변환
# =========================================================
def convert_long_format(df, default_category=None):
    item_col = find_column(df, COLUMN_ALIASES["item"])
    date_col = find_column(df, COLUMN_ALIASES["date"])
    value_col = find_column(df, COLUMN_ALIASES["value"])

    if not all([item_col, date_col, value_col]):
        return None

    unit_col = find_column(df, COLUMN_ALIASES["unit"])
    currency_col = find_column(df, COLUMN_ALIASES["currency"])
    related_col = find_column(df, COLUMN_ALIASES["related"])
    source_col = find_column(df, COLUMN_ALIASES["source"])
    note_col = find_column(df, COLUMN_ALIASES["note"])
    category_col = find_column(df, COLUMN_ALIASES["category"])

    result = pd.DataFrame({
        "item": df[item_col].astype(str).str.strip(),
        "date": df[date_col].apply(parse_month),
        "value": df[value_col].apply(clean_number),
        "unit": df[unit_col].astype(str).str.strip() if unit_col else "",
        "currency": df[currency_col].astype(str).str.strip() if currency_col else "",
        "related": df[related_col].astype(str).str.strip() if related_col else "",
        "source": df[source_col].astype(str).str.strip() if source_col else "",
        "note": df[note_col].astype(str).str.strip() if note_col else "",
        "category": (
            df[category_col].astype(str).str.strip()
            if category_col
            else default_category or ""
        ),
    })
    return result


def convert_wide_format(df, default_category=None):
    month_cols = [c for c in df.columns if looks_like_month_column(c)]
    if not month_cols:
        return None

    item_col = find_column(df, COLUMN_ALIASES["item"])
    if item_col is None:
        # 첫 번째 비월 컬럼을 품목으로 사용
        non_month_cols = [c for c in df.columns if c not in month_cols]
        if not non_month_cols:
            return None
        item_col = non_month_cols[0]

    unit_col = find_column(df, COLUMN_ALIASES["unit"])
    currency_col = find_column(df, COLUMN_ALIASES["currency"])
    related_col = find_column(df, COLUMN_ALIASES["related"])
    source_col = find_column(df, COLUMN_ALIASES["source"])
    note_col = find_column(df, COLUMN_ALIASES["note"])
    category_col = find_column(df, COLUMN_ALIASES["category"])

    id_vars = [item_col]
    for col in [unit_col, currency_col, related_col, source_col, note_col, category_col]:
        if col and col not in id_vars:
            id_vars.append(col)

    melted = df.melt(
        id_vars=id_vars,
        value_vars=month_cols,
        var_name="date_raw",
        value_name="value_raw",
    )

    result = pd.DataFrame({
        "item": melted[item_col].astype(str).str.strip(),
        "date": melted["date_raw"].apply(parse_month),
        "value": melted["value_raw"].apply(clean_number),
        "unit": melted[unit_col].astype(str).str.strip() if unit_col else "",
        "currency": melted[currency_col].astype(str).str.strip() if currency_col else "",
        "related": melted[related_col].astype(str).str.strip() if related_col else "",
        "source": melted[source_col].astype(str).str.strip() if source_col else "",
        "note": melted[note_col].astype(str).str.strip() if note_col else "",
        "category": (
            melted[category_col].astype(str).str.strip()
            if category_col
            else default_category or ""
        ),
    })
    return result


def tidy_dataframe(df, default_category=None):
    if df is None or df.empty:
        return pd.DataFrame(columns=[
            "item", "date", "value", "currency", "unit", "related", "source", "note", "category"
        ])

    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all").dropna(axis=1, how="all")

    result = convert_long_format(df, default_category)
    if result is None:
        result = convert_wide_format(df, default_category)

    if result is None:
        raise ValueError(
            "데이터 형식을 인식하지 못했습니다. "
            "세로형은 [품목, 날짜, 값], 가로형은 [품목, 2026-01, 2026-02...] 형태로 입력해 주세요."
        )

    for col in ["unit", "currency", "related", "source", "note", "category"]:
        result[col] = result[col].replace({"nan": "", "None": ""}).fillna("").astype(str).str.strip()

    result["item"] = result["item"].replace({"nan": "", "None": ""}).fillna("").astype(str).str.strip()
    split_items = result["item"].map(_split_item_note)
    result["item"] = split_items.map(lambda value: value[0])
    extracted_notes = split_items.map(lambda value: value[1])
    result["note"] = result["note"].where(result["note"] != "", extracted_notes)
    result["value"] = pd.to_numeric(result["value"], errors="coerce")
    result = result.dropna(subset=["date", "value"])
    result = result[result["item"] != ""]

    # 동일 품목·동일 월 중복 시 마지막 값 사용
    result = (
        result.sort_values(["item", "date"])
        .drop_duplicates(subset=["item", "date"], keep="last")
        .reset_index(drop=True)
    )
    result["ym"] = result["date"].map(lambda value: f"{value.year}년 {value.month}월")
    return result


@st.cache_data(show_spinner=False)
def load_excel(file_bytes):
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    sheet_names = xls.sheet_names

    raw_sheet = select_sheet(sheet_names, "원자재")
    fx_sheet = select_sheet(sheet_names, "환율")

    loaded = {}
    errors = []

    if raw_sheet:
        try:
            raw_source = pd.read_excel(io.BytesIO(file_bytes), sheet_name=raw_sheet)
            loaded["원자재"] = tidy_dataframe(raw_source, "원자재")
        except Exception as exc:
            errors.append(f"원자재 시트: {exc}")

    if fx_sheet:
        try:
            fx_source = pd.read_excel(io.BytesIO(file_bytes), sheet_name=fx_sheet)
            loaded["환율"] = tidy_dataframe(fx_source, "환율")
        except Exception as exc:
            errors.append(f"환율 시트: {exc}")

    # 원자재/환율 시트명이 없으면 첫 시트의 '구분' 열로 분리 시도
    if not loaded and sheet_names:
        first_df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_names[0])
        combined = tidy_dataframe(first_df)

        if "category" in combined.columns and combined["category"].str.strip().ne("").any():
            category_text = combined["category"].str.lower()
            raw_mask = category_text.str.contains("원자재|raw|commodity", regex=True)
            fx_mask = category_text.str.contains("환율|fx|exchange|currency", regex=True)

            if raw_mask.any():
                loaded["원자재"] = combined[raw_mask].copy()
            if fx_mask.any():
                loaded["환율"] = combined[fx_mask].copy()

        if not loaded:
            loaded["원자재"] = combined
            errors.append(
                "시트명을 찾지 못해 첫 번째 시트를 원자재 데이터로 불러왔습니다."
            )

    return loaded, errors, sheet_names


# =========================================================
# 연도별 업로드 양식 변환기
# =========================================================
_load_legacy_excel = load_excel
YEAR_SHEET_PATTERN = re.compile(r"^\s*(20\d{2})\s*년?\s*$")
FX_CODES = {"USD", "EUR", "JPY", "CNY", "GBP", "AUD", "CAD", "CHF"}


def _upload_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def _split_item_note(value):
    """품목명 끝의 괄호 문구를 품목명과 표시용 메모로 분리합니다."""
    text = _upload_text(value)
    matched = re.match(r"^(.*?)\s*\(([^()]*)\)\s*$", text)
    if not matched:
        return text, ""
    return matched.group(1).strip(), matched.group(2).strip()


def _currency_code(value):
    text = _upload_text(value).upper()
    matched = re.search(r"\b([A-Z]{3})\b", text)
    return matched.group(1) if matched else text


def _month_header(value, year):
    if pd.isna(value):
        return pd.NaT
    if isinstance(value, (pd.Timestamp, datetime)):
        return pd.Timestamp(value.year, value.month, 1)
    if hasattr(value, "year") and hasattr(value, "month"):
        return pd.Timestamp(int(value.year), int(value.month), 1)

    text = _upload_text(value)
    full = re.search(r"(20\d{2})\D+(1[0-2]|0?[1-9])", text)
    if full:
        return pd.Timestamp(int(full.group(1)), int(full.group(2)), 1)
    month = re.fullmatch(r"(1[0-2]|0?[1-9])\s*월?", text)
    if month:
        return pd.Timestamp(year, int(month.group(1)), 1)
    return pd.NaT


def _find_upload_header(raw):
    for row_no in range(min(len(raw), 20)):
        values = {_upload_text(value).lower() for value in raw.iloc[row_no].tolist()}
        if ({"구분", "원료항목", "항목"} & values) and ({"no.", "no", "번호"} & values):
            return row_no
    raise ValueError("'no.'와 '구분'이 있는 헤더 행을 찾지 못했습니다.")


def _read_year_sheet(xls, sheet_name, year):
    raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
    header_row = _find_upload_header(raw)
    header = raw.iloc[header_row].tolist()
    aliases = {
        "item": {"구분", "원료항목", "항목"},
        "related": {"관련 구매품", "관련구매품"},
        "currency": {"통화"},
        "unit": {"단위"},
        "note": {"비고", "메모", "참고"},
    }
    positions = {}
    for column_no, value in enumerate(header):
        label = re.sub(r"\s+", " ", _upload_text(value).lower())
        for name, candidates in aliases.items():
            if label in candidates and name not in positions:
                positions[name] = column_no

    month_columns = []
    for column_no, value in enumerate(header):
        month = _month_header(value, year)
        if pd.notna(month):
            month_columns.append((column_no, month))
    if "item" not in positions or not month_columns:
        raise ValueError(f"{sheet_name}: 항목 또는 월별 데이터 열을 찾지 못했습니다.")

    note_column = positions.get("note")
    if note_column is None:
        no_column = next(
            (
                column_no
                for column_no, value in enumerate(header)
                if _upload_text(value).lower() in {"no.", "no", "번호"}
            ),
            None,
        )
        if no_column is not None and no_column > 0:
            note_column = no_column - 1

    records = []
    for row_no in range(header_row + 1, len(raw)):
        row = raw.iloc[row_no]
        item = _upload_text(row.iloc[positions["item"]])
        if not item:
            continue
        item, item_note = _split_item_note(item)
        note = _upload_text(row.iloc[note_column]) if note_column is not None else ""
        note = note or item_note
        currency = _upload_text(row.iloc[positions["currency"]]) if "currency" in positions else ""
        category = "환율" if _currency_code(item) in FX_CODES and not currency else "원자재"
        unit = _upload_text(row.iloc[positions["unit"]]) if "unit" in positions else ""
        related = _upload_text(row.iloc[positions["related"]]) if "related" in positions else ""
        if category == "환율":
            item = _currency_code(item)
            currency = "KRW"
            unit = "100 JPY" if item == "JPY" else f"1 {item}"

        for column_no, month in month_columns:
            value = clean_number(row.iloc[column_no])
            if pd.isna(value):
                continue
            records.append({
                "item": item,
                "date": month,
                "value": value,
                "unit": unit,
                "currency": currency,
                "related": related,
                "source": "",
                "note": note,
                "category": category,
            })
    return pd.DataFrame(records)


def _read_source_sheet(xls):
    source_name = next((name for name in xls.sheet_names if _upload_text(name) == "출처"), None)
    if source_name is None:
        return {}
    raw = pd.read_excel(xls, sheet_name=source_name, header=None)
    source_map = {}
    section = ""
    for row_no in range(len(raw)):
        values = [_upload_text(value) for value in raw.iloc[row_no].tolist()]
        joined = " ".join(value for value in values if value)
        if "원자재" in joined and "no." not in joined.lower():
            section = "원자재"
        elif "환율" in joined and "no." not in joined.lower():
            section = "환율"
        lowered = [value.lower() for value in values]
        if not ({"no.", "no", "번호"} & set(lowered)) or "출처" not in lowered:
            continue
        item_col = lowered.index("원료항목") if "원료항목" in lowered else (
            lowered.index("통화") if "통화" in lowered else None
        )
        if item_col is None:
            continue
        source_col = lowered.index("출처")
        for data_row_no in range(row_no + 1, len(raw)):
            data = [_upload_text(value) for value in raw.iloc[data_row_no].tolist()]
            if any("▶" in value for value in data) or ({"no.", "no", "번호"} & {value.lower() for value in data}):
                break
            if item_col >= len(data) or not data[item_col]:
                continue
            if section == "환율":
                item = _currency_code(data[item_col])
            else:
                item, _ = _split_item_note(data[item_col])
            source_map[(section, item)] = data[source_col] if source_col < len(data) else ""
    return source_map


@st.cache_data(show_spinner=False)
def load_yearly_upload_excel(file_bytes):
    xls = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
    frames = []
    errors = []
    for sheet_name in xls.sheet_names:
        matched = YEAR_SHEET_PATTERN.match(_upload_text(sheet_name))
        if not matched:
            continue
        try:
            frame = _read_year_sheet(xls, sheet_name, int(matched.group(1)))
            if not frame.empty:
                frames.append(frame)
        except Exception as exc:
            errors.append(str(exc))

    if not frames:
        # 기존 세로형/가로형 파일도 계속 지원합니다.
        return _load_legacy_excel(file_bytes)

    combined = pd.concat(frames, ignore_index=True)
    source_map = _read_source_sheet(xls)
    combined["source"] = combined.apply(
        lambda row: source_map.get((row["category"], row["item"]), ""), axis=1
    )
    combined = (
        combined.sort_values(["category", "item", "date"])
        .drop_duplicates(["category", "item", "date"], keep="last")
        .reset_index(drop=True)
    )
    combined["ym"] = combined["date"].map(lambda value: f"{value.year}년 {value.month}월")
    loaded = {
        "원자재": combined[combined["category"] == "원자재"].copy(),
        "환율": combined[combined["category"] == "환율"].copy(),
    }
    return loaded, errors, xls.sheet_names


# 화면 구성은 기존 앱 그대로 두고 업로드 변환기만 새 양식 지원 버전으로 교체합니다.
load_excel = load_yearly_upload_excel


# =========================================================
# 분석 계산
# =========================================================
def filter_period(df, months_to_show):
    if df.empty:
        return df

    latest_month = df["date"].max()
    start_month = latest_month - pd.DateOffset(months=months_to_show - 1)
    return df[df["date"].between(start_month, latest_month)].copy()


def get_item_meta(df):
    if "currency" not in df.columns:
        df = df.copy()
        df["currency"] = ""
    if "note" not in df.columns:
        df = df.copy()
        df["note"] = ""
    meta = (
        df.sort_values("date")
        .groupby("item", as_index=False)
        .agg({
            "unit": "last",
            "currency": "last",
            "related": "last",
            "source": "last",
            "note": "last",
        })
    )
    return meta.set_index("item").to_dict("index")


def get_snapshot(df):
    rows = []
    for item, group in df.groupby("item"):
        group = group.sort_values("date")
        latest_row = group.iloc[-1]
        latest_date = latest_row["date"]
        latest_value = latest_row["value"]

        prev_date = latest_date - pd.DateOffset(months=1)
        prev_match = group[group["date"] == prev_date]
        prev_value = prev_match.iloc[-1]["value"] if not prev_match.empty else pd.NA

        six_month_date = latest_date - pd.DateOffset(months=6)
        six_month_match = group[group["date"] == six_month_date]
        six_month_value = (
            six_month_match.iloc[-1]["value"] if not six_month_match.empty else pd.NA
        )

        previous_year = latest_date.year - 1
        previous_year_values = group.loc[
            group["date"].dt.year == previous_year, "value"
        ].dropna()
        previous_year_average = (
            previous_year_values.mean() if not previous_year_values.empty else pd.NA
        )

        mom = (
            (latest_value / prev_value - 1) * 100
            if pd.notna(prev_value) and prev_value != 0
            else pd.NA
        )
        mom_difference = (
            latest_value - prev_value if pd.notna(prev_value) else pd.NA
        )
        six_month_difference = (
            latest_value - six_month_value if pd.notna(six_month_value) else pd.NA
        )
        six_month_change = (
            (latest_value / six_month_value - 1) * 100
            if pd.notna(six_month_value) and six_month_value != 0
            else pd.NA
        )
        previous_year_difference = (
            latest_value - previous_year_average
            if pd.notna(previous_year_average)
            else pd.NA
        )
        previous_year_change = (
            (latest_value / previous_year_average - 1) * 100
            if pd.notna(previous_year_average) and previous_year_average != 0
            else pd.NA
        )

        rows.append({
            "item": item,
            "latest_date": latest_date,
            "latest_value": latest_value,
            "prev_value": prev_value,
            "mom_difference": mom_difference,
            "mom": mom,
            "six_month_value": six_month_value,
            "six_month_difference": six_month_difference,
            "six_month_change": six_month_change,
            "previous_year": previous_year,
            "previous_year_average": previous_year_average,
            "previous_year_difference": previous_year_difference,
            "previous_year_change": previous_year_change,
        })

    return pd.DataFrame(rows)


def format_value(value, unit=""):
    if pd.isna(value):
        return "-"

    if abs(value) >= 1000:
        text = f"{value:,.2f}"
    else:
        text = f"{value:,.2f}"

    return f"{text} {unit}".strip()


def combined_unit(currency="", unit=""):
    currency = str(currency or "").strip()
    unit = str(unit or "").strip()
    if currency and unit:
        return f"{currency} / {unit}"
    return currency or unit


def format_delta(value):
    if pd.isna(value):
        return "-"
    if value > 0:
        return f"▲ {abs(value):.2f}%"
    if value < 0:
        return f"▼ {abs(value):.2f}%"
    return "• 0.00%"


def format_table_delta(value):
    if pd.isna(value):
        return "-"
    if value > 0:
        return f"🔴 ▲ {abs(value):.2f}%"
    if value < 0:
        return f"🔵 ▼ {abs(value):.2f}%"
    return "⚪ • 0.00%"


def format_signed_value(value, unit=""):
    if pd.isna(value):
        return "-"
    if value > 0:
        prefix = "▲"
    elif value < 0:
        prefix = "▼"
    else:
        prefix = "•"
    return f"{prefix} {abs(value):,.2f} {unit}".strip()


def colorize_change(value):
    text = str(value).strip()
    if text.startswith("▲"):
        css_class = "change-up"
    if text.startswith("▼"):
        css_class = "change-down"
    elif text.startswith("•"):
        css_class = "change-flat"
    elif not text.startswith("▲"):
        return html.escape(text)
    return f'<span class="{css_class}">{html.escape(text)}</span>'


def map_dataframe(frame, function):
    """Apply a cell formatter across old and new pandas versions."""
    if hasattr(frame, "map"):
        return frame.map(function)
    return frame.applymap(function)


# =========================================================
# 화면 렌더링
# =========================================================
def render_kpis(df, show_related=True):
    snapshot = get_snapshot(df)
    metadata = get_item_meta(df)
    cols = st.columns(4)

    for index, row in snapshot.iterrows():
        meta = metadata.get(row["item"], {})
        related = meta.get("related") or "-"
        unit = meta.get("unit") or ""
        value_unit = combined_unit(meta.get("currency"), unit)

        with cols[index % 4]:
            with st.container(border=True):
                st.markdown(f"**{row['item']}**")
                if show_related:
                    st.caption(f"사용 품목: {related}")
                if pd.isna(row["mom"]):
                    delta_badge = '<span class="metric-flat">전월 데이터 없음</span>'
                elif row["mom"] > 0:
                    delta_badge = (
                        f'<span class="metric-up">▲ 전월 대비 {abs(row["mom"]):.2f}%</span>'
                    )
                elif row["mom"] < 0:
                    delta_badge = (
                        f'<span class="metric-down">▼ 전월 대비 {abs(row["mom"]):.2f}%</span>'
                    )
                else:
                    delta_badge = '<span class="metric-flat">• 전월 대비 0.00%</span>'

                metric_html = (
                    '<div class="custom-metric">'
                    f'<div class="metric-month">{row["latest_date"].year}년 '
                    f'{row["latest_date"].month}월</div>'
                    f'<div class="metric-value">{html.escape(format_value(row["latest_value"], value_unit))}</div>'
                    f'{delta_badge}</div>'
                )
                st.markdown(
                    "<style>"
                    ".custom-metric{background:#f8f9fb;border:1px solid #e1e5eb;"
                    "border-radius:12px;padding:14px 14px 13px;margin-top:8px}"
                    ".metric-month{font-size:.82rem;color:#54616f;margin-bottom:4px}"
                    ".metric-value{font-size:1.42rem;color:#17233b;font-weight:750;"
                    "margin-bottom:8px;white-space:nowrap}"
                    ".metric-up,.metric-down,.metric-flat{display:inline-block;border-radius:999px;"
                    "padding:3px 8px;font-size:.83rem;font-weight:750}"
                    ".metric-up{color:#c62828;background:#ffebee}"
                    ".metric-down{color:#1565c0;background:#e3f2fd}"
                    ".metric-flat{color:#616161;background:#eeeeee}"
                    "</style>" + metric_html,
                    unsafe_allow_html=True,
                )


def render_auto_analysis(df, key):
    """업로드된 수치만 사용해 추세·변동성·주의 항목을 자동 요약합니다."""
    snapshot = get_snapshot(df)
    if snapshot.empty:
        st.info("자동 분석에 사용할 데이터가 없습니다.")
        return

    insights = []
    valid_mom = snapshot.dropna(subset=["mom"]).copy()
    if not valid_mom.empty:
        rise_count = int((valid_mom["mom"] > 0).sum())
        fall_count = int((valid_mom["mom"] < 0).sum())
        flat_count = int((valid_mom["mom"] == 0).sum())
        insights.append(
            f"최신 전월 대비 기준으로 **상승 {rise_count}개, 하락 {fall_count}개, "
            f"보합 {flat_count}개**입니다."
        )

        biggest_rise = valid_mom.loc[valid_mom["mom"].idxmax()]
        biggest_fall = valid_mom.loc[valid_mom["mom"].idxmin()]
        if biggest_rise["mom"] > 0:
            insights.append(
                f"가장 큰 상승 항목은 **{biggest_rise['item']} "
                f"(▲ {abs(biggest_rise['mom']):.2f}%)**입니다."
            )
        if biggest_fall["mom"] < 0:
            insights.append(
                f"가장 큰 하락 항목은 **{biggest_fall['item']} "
                f"(▼ {abs(biggest_fall['mom']):.2f}%)**입니다."
            )

    valid_six_month = snapshot.dropna(subset=["six_month_change"]).copy()
    if not valid_six_month.empty:
        strongest_six_month = valid_six_month.loc[
            valid_six_month["six_month_change"].abs().idxmax()
        ]
        direction = "상승" if strongest_six_month["six_month_change"] > 0 else "하락"
        insights.append(
            f"최근 6개월 변화가 가장 큰 항목은 **{strongest_six_month['item']}**이며, "
            f"6개월 전 대비 **{abs(strongest_six_month['six_month_change']):.2f}% "
            f"{direction}**했습니다."
        )

    volatility_rows = []
    for item, item_df in df.groupby("item"):
        recent_values = (
            item_df.sort_values("date")
            .drop_duplicates("date", keep="last")
            .tail(6)["value"]
        )
        returns = recent_values.pct_change(fill_method=None).dropna() * 100
        if len(returns) >= 2:
            volatility_rows.append((item, float(returns.std())))
    if volatility_rows:
        volatile_item, volatility = max(volatility_rows, key=lambda row: row[1])
        insights.append(
            f"최근 6개월 월간 변동성이 가장 큰 항목은 **{volatile_item}** "
            f"(변동성 {volatility:.2f}%)로, 가격 변화를 우선 점검할 필요가 있습니다."
        )

    if key == "환율":
        rising_fx = valid_mom[valid_mom["mom"] > 0] if not valid_mom.empty else valid_mom
        if not rising_fx.empty:
            fx_names = ", ".join(rising_fx.sort_values("mom", ascending=False)["item"].head(3))
            insights.append(
                f"**구매 관점:** {fx_names} 환율 상승은 해당 통화 결제 품목의 "
                "원화 환산 원가를 높일 수 있으므로 결제 시점과 환율 변동을 함께 확인하세요."
            )
        else:
            insights.append(
                "**구매 관점:** 주요 환율의 전월 상승 신호는 제한적이지만, "
                "결제 통화별 추세를 지속적으로 확인하는 것이 좋습니다."
            )
    else:
        watch = valid_mom[valid_mom["mom"] >= 5].sort_values("mom", ascending=False)
        if watch.empty:
            watch = valid_six_month[
                valid_six_month["six_month_change"] >= 10
            ].sort_values("six_month_change", ascending=False)
        if not watch.empty:
            watch_names = ", ".join(watch["item"].head(3))
            insights.append(
                f"**구매 관점:** **{watch_names}**은 최근 상승 폭이 상대적으로 커 "
                "견적·발주 시점과 재고 수준을 우선 점검하는 것이 좋습니다."
            )
        else:
            insights.append(
                "**구매 관점:** 현재 수치에서는 급격한 상승 경보가 제한적입니다. "
                "다만 환율과 원자재 가격의 동반 상승 여부를 계속 확인하세요."
            )

    latest_date = snapshot["latest_date"].max()
    with st.container(border=True):
        st.markdown("### 자동 분석 의견")
        st.caption(
            f"{latest_date.year}년 {latest_date.month}월 최신 수치 기준 · "
            "업로드 데이터의 증감률과 변동성을 자동 분석한 참고 의견입니다."
        )
        for insight in insights:
            st.markdown(f"- {insight}")
        st.caption(
            "※ 시장 뉴스나 계약 조건은 반영하지 않은 수치 기반 분석이므로 "
            "최종 구매 판단 시 공급사 정보와 함께 검토하세요."
        )


def build_integrated_analysis(datasets):
    rows = []
    for category in ["원자재", "환율"]:
        category_df = datasets.get(category)
        if category_df is None or category_df.empty:
            continue
        snapshot = get_snapshot(category_df)
        metadata = get_item_meta(category_df)
        category_latest_date = category_df["date"].max()
        for _, item_row in snapshot.iterrows():
            item = item_row["item"]
            item_values = (
                category_df[category_df["item"] == item]
                .sort_values("date")
                .drop_duplicates("date", keep="last")
                .tail(6)["value"]
            )
            returns = item_values.pct_change(fill_method=None).dropna() * 100
            volatility = float(returns.std()) if len(returns) >= 2 else 0.0
            mom = float(item_row["mom"]) if pd.notna(item_row["mom"]) else pd.NA
            six_month = (
                float(item_row["six_month_change"])
                if pd.notna(item_row["six_month_change"])
                else pd.NA
            )

            if item_row["latest_date"] < category_latest_date or pd.isna(mom):
                signal = "⚠️ 데이터 확인"
                signal_group = "데이터 확인"
            elif mom >= 1.5 and pd.notna(six_month) and six_month >= 5:
                signal = "🔴 상승 압력"
                signal_group = "상승 압력"
            elif mom <= -1.5 and pd.notna(six_month) and six_month <= -5:
                signal = "🔵 하락 완화"
                signal_group = "하락 완화"
            elif mom >= 1.5 and pd.notna(six_month) and six_month < 0:
                signal = "🟠 상승 전환"
                signal_group = "방향 전환"
            elif mom <= -1.5 and pd.notna(six_month) and six_month > 0:
                signal = "🟣 하락 전환"
                signal_group = "방향 전환"
            elif abs(mom) < 1.5:
                signal = "⚪ 보합"
                signal_group = "보합"
            elif mom > 0:
                signal = "🔺 상승 관찰"
                signal_group = "상승 관찰"
            else:
                signal = "🔻 하락 관찰"
                signal_group = "하락 관찰"

            meta = metadata.get(item, {})
            related = meta.get("related") or "-"
            if signal_group == "데이터 확인":
                action = "최신 월 데이터 입력 여부와 업데이트 주기를 확인"
            elif category == "환율":
                if signal_group in ["상승 압력", "상승 관찰", "방향 전환"] and mom > 0:
                    action = "해당 통화 결제 예정액·결제일·적용 환율 기준일을 확인"
                elif mom < 0:
                    action = "환율 하락 효과가 최근 견적에 반영됐는지 확인"
                else:
                    action = "정기적으로 결제 통화의 월별 방향을 확인"
            elif signal_group == "상승 압력":
                action = f"{related} 관련 최근 견적·발주 예정 물량·재고 수준을 확인"
            elif signal_group == "하락 완화":
                action = f"{related} 관련 공급사 단가 인하 협의 가능 여부를 확인"
            elif signal_group == "방향 전환":
                action = "일시적 반등·조정인지 다음 월 데이터와 공급사 견적을 함께 확인"
            elif mom > 0:
                action = "최근 견적에 상승분이 반영됐는지 확인"
            elif mom < 0:
                action = "가격 하락분의 구매단가 반영 가능 여부를 확인"
            else:
                action = "현재 계약단가와 시장가격의 차이를 정기 확인"

            change_score = (
                (abs(float(mom)) if pd.notna(mom) else 0) * 1.5
                + (abs(float(six_month)) if pd.notna(six_month) else 0) * 0.5
            )

            rows.append(
                {
                    "구분": category,
                    "품목": item,
                    "최근 기준월": item_row["latest_date"],
                    "전월 증감률": mom,
                    "6개월 증감률": six_month,
                    "최근 변동성": volatility,
                    "가격 방향 신호": signal,
                    "신호 구분": signal_group,
                    "확인 필요사항": action,
                    "_change_score": change_score,
                    "_currency": meta.get("currency") or "",
                    "_related": related,
                }
            )
    return pd.DataFrame(rows)


def render_purchase_trend(datasets, analysis):
    options = []
    data_lookup = {}
    for category in ["원자재", "환율"]:
        category_df = datasets.get(category)
        if category_df is None or category_df.empty:
            continue
        for item in sorted(category_df["item"].unique()):
            label = f"{category} · {item}"
            options.append(label)
            data_lookup[label] = category_df[category_df["item"] == item].copy()

    default_labels = [
        f"{row['구분']} · {row['품목']}"
        for _, row in analysis.sort_values("_change_score", ascending=False).head(3).iterrows()
    ]
    selected = st.multiselect(
        "추세 비교 품목",
        options=options,
        default=[label for label in default_labels if label in options],
        key="purchase_insight_trend_items",
    )
    if not selected:
        st.info("추세를 비교할 품목을 1개 이상 선택해 주세요.")
        return

    fig = go.Figure()
    colors = ["#1565C0", "#C62828", "#2E7D32", "#8E24AA", "#EF6C00", "#00838F"]
    for index, label in enumerate(selected):
        item_df = (
            data_lookup[label]
            .sort_values("date")
            .drop_duplicates("date", keep="last")
            .tail(12)
        )
        if item_df.empty or item_df.iloc[0]["value"] == 0:
            continue
        color = colors[index % len(colors)]
        indexed = item_df["value"] / item_df.iloc[0]["value"] * 100
        moving_average = indexed.rolling(3, min_periods=2).mean()
        fig.add_trace(
            go.Scatter(
                x=item_df["date"],
                y=indexed,
                mode="lines+markers",
                name=label,
                line={"color": color, "width": 2.5},
                marker={"size": 6},
                hovertemplate="%{x|%Y년 %m월}<br>지수 %{y:.1f}<extra>%{fullData.name}</extra>",
            )
        )
        fig.add_trace(
            go.Scatter(
                x=item_df["date"],
                y=moving_average,
                mode="lines",
                name=f"{label} · 3개월 평균",
                line={"color": color, "width": 2, "dash": "dot"},
                hovertemplate="%{x|%Y년 %m월}<br>3개월 평균 %{y:.1f}<extra>%{fullData.name}</extra>",
            )
        )
    fig.add_hline(y=100, line_dash="dash", line_color="#9E9E9E")
    fig.update_layout(
        height=500,
        margin={"l": 20, "r": 20, "t": 20, "b": 20},
        hovermode="x unified",
        legend={"orientation": "h", "y": 1.02, "x": 0},
        xaxis={"title": "", "tickformat": "%Y년 %m월"},
        yaxis={"title": "지수 (조회 시작월=100)"},
    )
    st.plotly_chart(fig, use_container_width=True)
    st.caption("실선은 월별 실제 움직임, 점선은 단기 방향을 보기 위한 3개월 이동평균입니다.")


def render_ai_insights(datasets):
    analysis = build_integrated_analysis(datasets)
    if analysis.empty:
        st.info("구매 분석 인사이트에 사용할 데이터가 없습니다.")
        return

    latest_date = analysis["최근 기준월"].max()
    st.subheader("구매 분석 인사이트")
    st.caption(
        f"{latest_date.year}년 {latest_date.month}월 최신 데이터 기준 · "
        "원자재와 환율의 전월·6개월 방향을 구매 확인 업무와 연결한 수치 기반 인사이트입니다."
    )

    signal_counts = analysis["신호 구분"].value_counts()
    metric_columns = st.columns(5)
    metric_columns[0].metric("분석 항목", f"{len(analysis)}개")
    metric_columns[1].metric("상승 압력", f"{signal_counts.get('상승 압력', 0)}개")
    metric_columns[2].metric("하락 완화", f"{signal_counts.get('하락 완화', 0)}개")
    metric_columns[3].metric("방향 전환", f"{signal_counts.get('방향 전환', 0)}개")
    metric_columns[4].metric("데이터 확인", f"{signal_counts.get('데이터 확인', 0)}개")

    st.markdown("### 핵심 변화")
    top_items = analysis.sort_values("_change_score", ascending=False).head(3)
    top_columns = st.columns(3)
    for column, (_, row) in zip(top_columns, top_items.iterrows()):
        with column:
            with st.container(border=True):
                st.caption(row["구분"])
                st.markdown(f"#### {row['품목']}")
                st.markdown(f"**{row['가격 방향 신호']}**")
                mom_text = "-" if pd.isna(row["전월 증감률"]) else f"{row['전월 증감률']:+.2f}%"
                six_text = "-" if pd.isna(row["6개월 증감률"]) else f"{row['6개월 증감률']:+.2f}%"
                st.write(f"전월 {mom_text} · 6개월 {six_text}")
                st.caption(row["확인 필요사항"])

    st.markdown("### 원자재·환율 복합 영향")
    raw_analysis = analysis[analysis["구분"] == "원자재"]
    fx_analysis = analysis[analysis["구분"] == "환율"]
    compound_rows = []
    for _, raw_row in raw_analysis.iterrows():
        currency = str(raw_row["_currency"]).upper()
        raw_change = raw_row["전월 증감률"]
        if currency == "KRW":
            fx_text = "환율 직접 영향 없음"
            compound_signal = "원자재 가격 방향 중심으로 확인"
        else:
            fx_match = fx_analysis[fx_analysis["품목"].astype(str).str.upper() == currency]
            if fx_match.empty or pd.isna(raw_change):
                fx_text = "환율 데이터 없음"
                compound_signal = "복합 영향 판단 보류"
            else:
                fx_change = fx_match.iloc[0]["전월 증감률"]
                fx_text = "-" if pd.isna(fx_change) else f"{fx_change:+.2f}%"
                if pd.isna(fx_change):
                    compound_signal = "복합 영향 판단 보류"
                elif raw_change > 0 and fx_change > 0:
                    compound_signal = "원자재·환율 동반 상승 압력"
                elif raw_change > 0 and fx_change < 0:
                    compound_signal = "환율 하락이 원자재 상승 일부 상쇄"
                elif raw_change < 0 and fx_change > 0:
                    compound_signal = "환율 상승이 원자재 하락 효과 일부 축소"
                elif raw_change < 0 and fx_change < 0:
                    compound_signal = "원자재·환율 동반 하락 완화"
                else:
                    compound_signal = "한쪽 요인 보합"
        compound_rows.append(
            {
                "원자재": raw_row["품목"],
                "관련 구매품": raw_row["_related"],
                "통화": currency,
                "원자재 전월": "-" if pd.isna(raw_change) else f"{raw_change:+.2f}%",
                "해당 환율 전월": fx_text,
                "복합 방향": compound_signal,
            }
        )
    if compound_rows:
        st.dataframe(pd.DataFrame(compound_rows), hide_index=True, use_container_width=True)
    else:
        st.info("원자재와 환율을 함께 비교할 데이터가 없습니다.")

    st.markdown("### 확인 필요사항")
    action_rows = analysis.sort_values("_change_score", ascending=False)[
        [
            "구분",
            "품목",
            "가격 방향 신호",
            "전월 증감률",
            "6개월 증감률",
            "확인 필요사항",
        ]
    ].copy()
    for column in ["전월 증감률", "6개월 증감률"]:
        action_rows[column] = action_rows[column].map(
            lambda value: "-" if pd.isna(value) else f"{value:+.2f}%"
        )
    st.dataframe(
        action_rows,
        hide_index=True,
        use_container_width=True,
    )

    st.markdown("### 주요 품목 추세")
    render_purchase_trend(datasets, analysis)

    st.caption(
        "※ 가격 방향 신호는 업로드된 전월·6개월 수치만으로 계산합니다. "
        "구매량·재고·계약조건·가격 반영 시차·시장뉴스는 포함하지 않으므로 "
        "최종 구매 판단이 아닌 확인 업무를 위한 참고자료로 사용하세요."
    )


def build_procurement_metrics(datasets):
    """구매 검토에 필요한 객관적인 월별 지표를 품목 단위로 계산합니다."""
    rows = []
    for category in ["원자재", "환율"]:
        category_df = datasets.get(category)
        if category_df is None or category_df.empty:
            continue
        snapshot = get_snapshot(category_df)
        metadata = get_item_meta(category_df)
        category_latest_date = category_df["date"].max()
        for _, snap in snapshot.iterrows():
            item = snap["item"]
            history = (
                category_df[category_df["item"] == item]
                .sort_values("date")
                .drop_duplicates("date", keep="last")
            )
            latest_value = float(snap["latest_value"])
            average_3m = float(history.tail(3)["value"].mean())
            versus_3m = (latest_value / average_3m - 1) * 100 if average_3m else pd.NA
            range_values = history.tail(12)["value"]
            range_min = float(range_values.min())
            range_max = float(range_values.max())
            range_position = (
                (latest_value - range_min) / (range_max - range_min) * 100
                if range_max != range_min
                else 50.0
            )
            recent_returns = history.tail(6)["value"].pct_change(fill_method=None).dropna() * 100
            volatility = float(recent_returns.std()) if len(recent_returns) >= 2 else 0.0
            meta = metadata.get(item, {})
            mom = float(snap["mom"]) if pd.notna(snap["mom"]) else pd.NA
            six_month = (
                float(snap["six_month_change"])
                if pd.notna(snap["six_month_change"])
                else pd.NA
            )
            previous_year_change = (
                float(snap["previous_year_change"])
                if pd.notna(snap["previous_year_change"])
                else pd.NA
            )
            rows.append(
                {
                    "구분": category,
                    "품목": item,
                    "관련 구매품": meta.get("related") or "-",
                    "비고": meta.get("note") or "",
                    "통화": meta.get("currency") or "-",
                    "단위": meta.get("unit") or "-",
                    "최근 기준월": snap["latest_date"],
                    "최신값": latest_value,
                    "전월 증감률": mom,
                    "3개월 평균 대비": versus_3m,
                    "6개월 증감률": six_month,
                    "전년 평균 대비": previous_year_change,
                    "12개월 범위 위치": range_position,
                    "6개월 변동성": volatility,
                    "데이터 상태": "업데이트 필요" if snap["latest_date"] < category_latest_date else "최신",
                    "_change_score": (
                        (abs(float(mom)) if pd.notna(mom) else 0) * 1.5
                        + (abs(float(six_month)) if pd.notna(six_month) else 0) * 0.5
                    ),
                    "_currency": meta.get("currency") or "",
                    "_is_stale": bool(snap["latest_date"] < category_latest_date),
                }
            )
    return pd.DataFrame(rows)


def build_currency_impact(analysis):
    raw_analysis = analysis[analysis["구분"] == "원자재"]
    fx_analysis = analysis[analysis["구분"] == "환율"]
    rows = []
    for _, raw in raw_analysis.iterrows():
        currency = str(raw["_currency"]).upper()
        raw_change = raw["전월 증감률"]
        fx_change = pd.NA
        if currency == "KRW":
            impact = "환율 직접 영향 없음"
        else:
            fx_match = fx_analysis[fx_analysis["품목"].astype(str).str.upper() == currency]
            if fx_match.empty or pd.isna(raw_change):
                impact = "판단 데이터 부족"
            else:
                fx_change = fx_match.iloc[0]["전월 증감률"]
                if pd.isna(fx_change):
                    impact = "판단 데이터 부족"
                elif raw_change > 0 and fx_change > 0:
                    impact = "원자재·환율 동반 상승"
                elif raw_change > 0 and fx_change < 0:
                    impact = "환율 하락이 원자재 상승 일부 상쇄"
                elif raw_change < 0 and fx_change > 0:
                    impact = "환율 상승이 원자재 하락 효과 일부 축소"
                elif raw_change < 0 and fx_change < 0:
                    impact = "원자재·환율 동반 하락"
                else:
                    impact = "한쪽 요인 보합"
        rows.append(
            {
                "원자재": raw["품목"],
                "관련 구매품": raw["관련 구매품"],
                "통화": currency,
                "원자재 전월 증감률": raw_change,
                "해당 환율 전월 증감률": fx_change,
                "동반 영향": impact,
            }
        )
    return pd.DataFrame(rows)


def render_procurement_check_list(title, frame, mode):
    color = {"rise": "red", "decline": "blue", "stale": "gray"}[mode]
    rows_html = []
    for _, row in frame.head(4).iterrows():
        if mode == "stale":
            change = f"{row['최근 기준월'].year}.{row['최근 기준월'].month:02d} 기준"
        else:
            value = row["전월 증감률"]
            if pd.isna(value):
                value = row["3개월 평균 대비"]
                prefix = "3개월 평균 대비 "
            else:
                prefix = "전월 대비 "
            arrow = "▲" if value > 0 else "▼" if value < 0 else "―"
            change = f"{prefix}{arrow} {abs(value):.1f}%"
        rows_html.append(
            '<div class="insight-row">'
            '<div class="insight-row-top">'
            f'<span class="insight-item-name">{html.escape(str(row["품목"]))}</span>'
            f'<span class="insight-change {color}">{html.escape(change)}</span>'
            '</div>'
            f'<div class="insight-related">{html.escape(str(row["관련 구매품"]))}</div>'
            '</div>'
        )
    if frame.empty:
        rows_html.append('<div class="insight-related">해당 품목이 없습니다.</div>')
    elif len(frame) > 4:
        rows_html.append(
            f'<div class="insight-related" style="margin-top:10px;">외 {len(frame) - 4}개 품목</div>'
        )
    st.markdown(
        f'<div class="insight-panel {color}">'
        f'<div class="insight-panel-title">{html.escape(title)}</div>'
        f'{"".join(rows_html)}</div>',
        unsafe_allow_html=True,
    )


def render_insight_kpi(label, count, color):
    st.markdown(
        f'<div class="insight-kpi {color}">'
        f'<div class="insight-kpi-label">{html.escape(label)}</div>'
        f'<div class="insight-kpi-value">{count}개</div>'
        '</div>',
        unsafe_allow_html=True,
    )


def render_procurement_summary(price_rise, price_decline, joint_rise, stale_items):
    rise_names = ", ".join(price_rise["품목"].head(3).astype(str)) or "해당 없음"
    decline_names = ", ".join(price_decline["품목"].head(3).astype(str)) or "해당 없음"
    if len(joint_rise) > 0:
        tone = "red"
        headline = "원가 상승 부담을 우선 확인할 시점입니다"
    elif len(price_rise) > len(price_decline):
        tone = "red"
        headline = "가격 상승 품목의 발주 조건을 우선 확인하세요"
    elif len(price_decline) > len(price_rise):
        tone = "blue"
        headline = "공급사 단가 인하 협의를 검토하기 좋은 구간입니다"
    elif len(price_rise) or len(price_decline):
        tone = "orange"
        headline = "품목별 상승과 하락이 혼재된 구간입니다"
    else:
        tone = "green"
        headline = "현재 확인되는 가격 변동은 제한적입니다"

    sentences = [
        f"상승 확인 {len(price_rise)}개, 인하 검토 {len(price_decline)}개입니다.",
        f"상승 품목({rise_names})은 최신 견적과 발주 시점을 확인하고, "
        f"하락 품목({decline_names})은 공급사 단가 반영 여부를 점검하세요.",
    ]
    if len(joint_rise):
        joint_names = ", ".join(joint_rise["원자재"].head(3).astype(str))
        sentences.append(f"{joint_names}은 원자재와 환율이 함께 상승해 외화 구매 부담 확인이 필요합니다.")
    if len(stale_items):
        sentences.append(f"최신 데이터가 없는 {len(stale_items)}개 항목은 판단 전에 업데이트가 필요합니다.")

    st.markdown(
        f'<div class="insight-summary {tone}">'
        f'<div class="insight-summary-title">구매팀 종합 요약 · {html.escape(headline)}</div>'
        f'<div class="insight-summary-text">{html.escape(" ".join(sentences))}</div>'
        '</div>',
        unsafe_allow_html=True,
    )


def mean_percent(series):
    values = pd.to_numeric(series, errors="coerce").dropna()
    return float(values.mean()) if not values.empty else pd.NA


def trend_label(value):
    if pd.isna(value):
        return "데이터 없음"
    if value >= 1:
        return "상승 흐름"
    if value <= -1:
        return "하락 흐름"
    return "보합 흐름"


def trend_color(value):
    if pd.isna(value) or abs(value) < 1:
        return "gray"
    return "red" if value > 0 else "blue"


def render_trend_kpi(label, value, note, color):
    value_text = "-" if pd.isna(value) else f"{value:+.1f}%"
    st.markdown(
        f'<div class="insight-kpi {color}">'
        f'<div class="insight-kpi-label">{html.escape(label)}</div>'
        f'<div class="insight-kpi-value">{html.escape(value_text)}</div>'
        f'<div class="insight-related">{html.escape(note)}</div>'
        '</div>',
        unsafe_allow_html=True,
    )


def render_objective_trend_summary(raw_analysis, fx_analysis):
    raw_mom = mean_percent(raw_analysis["전월 증감률"])
    raw_six = mean_percent(raw_analysis["6개월 증감률"])
    fx_mom = mean_percent(fx_analysis["전월 증감률"])
    valid_mom = pd.to_numeric(raw_analysis["전월 증감률"], errors="coerce").dropna()
    rise_share = float((valid_mom > 0).mean() * 100) if not valid_mom.empty else pd.NA

    short_label = trend_label(raw_mom)
    medium_label = trend_label(raw_six)
    if short_label == medium_label:
        headline = f"원자재 가격은 단기·중기 모두 {short_label}입니다"
    elif short_label == "하락 흐름" and medium_label == "상승 흐름":
        headline = "중기 상승 이후 최근에는 하락 조정 흐름입니다"
    elif short_label == "상승 흐름" and medium_label == "하락 흐름":
        headline = "중기 하락 이후 최근에는 반등 흐름입니다"
    else:
        headline = f"단기는 {short_label}, 중기는 {medium_label}입니다"

    if pd.notna(raw_mom) and raw_mom >= 1:
        tone = "red"
    elif pd.notna(raw_mom) and raw_mom <= -1:
        tone = "blue"
    elif pd.notna(raw_mom):
        tone = "green"
    else:
        tone = "orange"

    raw_mom_text = "-" if pd.isna(raw_mom) else f"{raw_mom:+.1f}%"
    raw_six_text = "-" if pd.isna(raw_six) else f"{raw_six:+.1f}%"
    rise_share_text = "-" if pd.isna(rise_share) else f"{rise_share:.0f}%"
    fx_text = "-" if pd.isna(fx_mom) else f"{fx_mom:+.1f}%"
    description = (
        f"원자재 품목의 전월 증감률 단순 평균은 {raw_mom_text}, 6개월 증감률 평균은 {raw_six_text}입니다. "
        f"전월 대비 상승한 품목 비중은 {rise_share_text}이며, 주요 환율의 전월 증감률 평균은 {fx_text}입니다."
    )
    st.markdown(
        f'<div class="insight-summary {tone}">'
        f'<div class="insight-summary-title">전체 추세 요약 · {html.escape(headline)}</div>'
        f'<div class="insight-summary-text">{html.escape(description)}</div>'
        '</div>',
        unsafe_allow_html=True,
    )
    return raw_mom, raw_six, rise_share, fx_mom


def render_monthly_change_chart(raw_analysis):
    chart_df = raw_analysis[["품목", "전월 증감률"]].dropna().sort_values("전월 증감률")
    if chart_df.empty:
        st.info("전월 증감률을 비교할 데이터가 없습니다.")
        return
    colors = ["#d32f2f" if value > 0 else "#1976d2" if value < 0 else "#98a2b3"
              for value in chart_df["전월 증감률"]]
    fig = go.Figure(
        go.Bar(
            x=chart_df["전월 증감률"],
            y=chart_df["품목"],
            orientation="h",
            marker_color=colors,
            text=chart_df["전월 증감률"].map(lambda value: f"{value:+.1f}%"),
            textposition="outside",
            hovertemplate="%{y}<br>전월 대비 %{x:+.2f}%<extra></extra>",
        )
    )
    fig.add_vline(x=0, line_color="#667085", line_width=1)
    fig.update_layout(
        height=max(320, 55 * len(chart_df)),
        margin={"l": 20, "r": 55, "t": 10, "b": 35},
        xaxis={"title": "전월 대비 증감률", "ticksuffix": "%", "zeroline": False},
        yaxis={"title": ""},
        showlegend=False,
    )
    st.plotly_chart(fig, use_container_width=True)
    st.caption("빨간색은 전월 대비 상승, 파란색은 하락을 의미합니다.")


def render_ai_insights(datasets):
    analysis = build_procurement_metrics(datasets)
    if analysis.empty:
        st.info("구매 데이터 인사이트에 사용할 데이터가 없습니다.")
        return

    latest_date = analysis["최근 기준월"].max()
    raw_analysis = analysis[analysis["구분"] == "원자재"].copy()
    rise_condition = (
        (raw_analysis["전월 증감률"] >= 2)
        | (raw_analysis["3개월 평균 대비"] >= 2)
        | (raw_analysis["6개월 증감률"] >= 10)
    )
    decline_condition = (
        (raw_analysis["전월 증감률"] <= -2)
        | (raw_analysis["3개월 평균 대비"] <= -2)
        | (raw_analysis["6개월 증감률"] <= -5)
    )
    short_term_direction = (
        raw_analysis["전월 증감률"].fillna(0)
        + raw_analysis["3개월 평균 대비"].fillna(0)
    )
    price_rise = raw_analysis[
        rise_condition & (~decline_condition | (short_term_direction >= 0))
    ]
    price_decline = raw_analysis[
        decline_condition & (~rise_condition | (short_term_direction < 0))
    ]
    stale_items = analysis[analysis["_is_stale"]]
    compound_df = build_currency_impact(analysis)
    joint_rise = compound_df[compound_df["동반 영향"] == "원자재·환율 동반 상승"]

    st.subheader("구매 데이터 인사이트")
    st.caption(
        f"{latest_date.year}년 {latest_date.month}월 최신 데이터 기준 · "
        "업로드된 가격과 환율만으로 계산한 구매 검토용 객관 지표입니다."
    )

    summary_tab, detail_tab, impact_tab, trend_tab = st.tabs(
        ["핵심 요약", "품목별 지표", "원자재·환율 영향", "추세 비교"]
    )

    with summary_tab:
        raw_mom, raw_six, rise_share, fx_mom = render_objective_trend_summary(
            raw_analysis, analysis[analysis["구분"] == "환율"]
        )
        columns = st.columns(4)
        with columns[0]:
            render_trend_kpi(
                "원자재 전월 추세",
                raw_mom,
                trend_label(raw_mom),
                trend_color(raw_mom),
            )
        with columns[1]:
            render_trend_kpi(
                "원자재 6개월 추세",
                raw_six,
                trend_label(raw_six),
                trend_color(raw_six),
            )
        with columns[2]:
            render_trend_kpi(
                "상승 품목 비중",
                rise_share,
                "전월 대비 상승한 원자재 비중",
                "orange",
            )
        with columns[3]:
            render_trend_kpi(
                "환율 전월 추세",
                fx_mom,
                trend_label(fx_mom),
                trend_color(fx_mom),
            )

        st.markdown("### 원자재 품목별 전월 추세")
        render_monthly_change_chart(raw_analysis)
        if len(stale_items):
            stale_names = ", ".join(stale_items["품목"].astype(str))
            st.warning(f"최신 기준월이 다른 항목: {stale_names}")
        st.caption(
            "※ 평균은 품목별 증감률을 동일 비중으로 계산한 단순 평균이며, 구매량에 따른 가중치는 적용하지 않았습니다."
        )

    with detail_tab:
        category_filter = st.radio(
            "조회 구분",
            options=["전체", "원자재", "환율"],
            horizontal=True,
            key="insight_category_filter",
        )
        detail = analysis if category_filter == "전체" else analysis[analysis["구분"] == category_filter]
        detail = detail[
            [
                "구분", "품목", "관련 구매품", "통화", "단위", "최근 기준월",
                "전월 증감률", "3개월 평균 대비", "6개월 증감률", "전년 평균 대비",
                "12개월 범위 위치", "6개월 변동성", "데이터 상태",
            ]
        ].copy()
        detail["최근 기준월"] = detail["최근 기준월"].map(
            lambda value: f"{value.year}년 {value.month}월"
        )
        st.dataframe(
            detail,
            hide_index=True,
            use_container_width=True,
            column_config={
                "전월 증감률": st.column_config.NumberColumn(format="%+.2f%%"),
                "3개월 평균 대비": st.column_config.NumberColumn(format="%+.2f%%"),
                "6개월 증감률": st.column_config.NumberColumn(format="%+.2f%%"),
                "전년 평균 대비": st.column_config.NumberColumn(format="%+.2f%%"),
                "12개월 범위 위치": st.column_config.ProgressColumn(
                    help="최근 12개월 최저값을 0, 최고값을 100으로 본 현재 위치",
                    min_value=0,
                    max_value=100,
                    format="%.0f%%",
                ),
                "6개월 변동성": st.column_config.NumberColumn(
                    help="최근 6개월 월간 증감률의 표준편차",
                    format="%.2f%%",
                ),
            },
        )
        st.caption(
            "12개월 범위 위치가 100%에 가까울수록 최근 1년 고점 부근입니다. "
            "변동성은 상승·하락 방향이 아니라 월별 흔들림의 크기를 뜻합니다."
        )

    with impact_tab:
        st.markdown("### 결제 통화까지 함께 본 전월 변화")
        if compound_df.empty:
            st.info("원자재와 환율을 함께 비교할 데이터가 없습니다.")
        else:
            st.dataframe(
                compound_df,
                hide_index=True,
                use_container_width=True,
                column_config={
                    "원자재 전월 증감률": st.column_config.NumberColumn(format="%+.2f%%"),
                    "해당 환율 전월 증감률": st.column_config.NumberColumn(format="%+.2f%%"),
                },
            )
            st.caption(
                "외화 결제 품목은 원자재 가격과 해당 통화 환율이 함께 오를 때 원화 구매 부담이 커질 수 있습니다. "
                "KRW 품목은 환율 직접 영향을 표시하지 않습니다."
            )

    with trend_tab:
        st.markdown("### 품목별 상대 추세")
        render_purchase_trend(datasets, analysis)

    st.caption(
        "※ 모든 결과는 업로드된 월별 수치의 산술 계산입니다. 구매량·재고·계약조건·가격 반영 시차·시장뉴스는 "
        "포함하지 않으므로 최종 구매 결정이 아닌 확인 항목 선별에 사용하세요."
    )


def classify_price_trend(mom, six_month):
    if pd.isna(mom) or pd.isna(six_month):
        return "데이터 부족"
    if mom >= 1 and six_month >= 1:
        return "지속 상승"
    if mom <= -1 and six_month <= -1:
        return "지속 하락"
    if mom <= -1 and six_month >= 1:
        return "상승 후 조정"
    if mom >= 1 and six_month <= -1:
        return "하락 후 반등"
    return "보합·혼조"


def join_item_names(frame):
    if frame.empty:
        return "해당 없음"
    return ", ".join(frame["품목"].astype(str))


def render_trend_group_card(title, frame, color, note):
    st.markdown(
        f'<div class="trend-group-card {color}">'
        f'<div class="trend-group-title">{html.escape(title)}</div>'
        f'<div class="trend-group-items">{html.escape(join_item_names(frame))}</div>'
        f'<div class="trend-group-note">{html.escape(note)}</div>'
        '</div>',
        unsafe_allow_html=True,
    )


def build_monthly_trend_narrative(raw_analysis, fx_analysis, compound_df):
    grouped = raw_analysis.copy()
    grouped["추세 구분"] = grouped.apply(
        lambda row: classify_price_trend(row["전월 증감률"], row["6개월 증감률"]),
        axis=1,
    )
    rising = grouped[grouped["추세 구분"].isin(["지속 상승", "하락 후 반등"])]
    falling = grouped[grouped["추세 구분"].isin(["지속 하락", "상승 후 조정"])]
    stable = grouped[grouped["추세 구분"] == "보합·혼조"]

    if not rising.empty and not falling.empty:
        opening = "이번 달 원자재 가격은 하나의 방향보다 품목별 차별화가 나타났습니다."
    elif not rising.empty:
        opening = "이번 달 원자재 가격은 상승 또는 반등 흐름이 상대적으로 두드러졌습니다."
    elif not falling.empty:
        opening = "이번 달 원자재 가격은 하락 또는 조정 흐름이 상대적으로 두드러졌습니다."
    else:
        opening = "이번 달 원자재 가격은 뚜렷한 한 방향 없이 보합·혼조 흐름을 보였습니다."

    sentences = [opening]
    if not rising.empty:
        sentences.append(f"상승·반등 흐름은 {join_item_names(rising)}에서 확인됩니다.")
    if not falling.empty:
        sentences.append(f"하락·조정 흐름은 {join_item_names(falling)}에서 나타났습니다.")
    if not stable.empty:
        sentences.append(f"{join_item_names(stable)}은 상대적으로 방향성이 제한적이었습니다.")

    valid_change = raw_analysis.dropna(subset=["전월 증감률"])
    if not valid_change.empty:
        largest = valid_change.loc[valid_change["전월 증감률"].abs().idxmax()]
        direction = "상승" if largest["전월 증감률"] > 0 else "하락"
        sentences.append(
            f"전월 대비 변동 폭이 가장 큰 품목은 {largest['품목']}으로 "
            f"{abs(largest['전월 증감률']):.1f}% {direction}했습니다."
        )

    if not compound_df.empty:
        amplified = compound_df[
            compound_df["동반 영향"].isin(
                ["원자재·환율 동반 상승", "환율 상승이 원자재 하락 효과 일부 축소"]
            )
        ]
        offset = compound_df[
            compound_df["동반 영향"] == "환율 하락이 원자재 상승 일부 상쇄"
        ]
        if not amplified.empty:
            names = ", ".join(amplified["원자재"].astype(str))
            sentences.append(f"{names}은 환율 방향이 원화 환산 가격 부담을 확대하거나 하락 효과를 줄이는 방향이었습니다.")
        if not offset.empty:
            names = ", ".join(offset["원자재"].astype(str))
            sentences.append(f"{names}은 환율 하락이 원자재 가격 상승을 일부 상쇄하는 방향이었습니다.")

    return " ".join(sentences), grouped


def render_monthly_market_report(datasets):
    analysis = build_procurement_metrics(datasets)
    if analysis.empty:
        st.info("월간 동향을 분석할 데이터가 없습니다.")
        return

    latest_date = analysis["최근 기준월"].max()
    raw_analysis = analysis[analysis["구분"] == "원자재"].copy()
    fx_analysis = analysis[analysis["구분"] == "환율"].copy()
    stale_items = analysis[analysis["_is_stale"]]
    compound_df = build_currency_impact(analysis)

    st.subheader(f"{latest_date.year}년 {latest_date.month}월 원자재 가격 동향")
    st.caption("업로드된 월별 원자재 가격과 환율의 실제 변화만 분석한 월간 동향입니다.")

    st.markdown("### 1. 이달의 종합 동향")
    narrative, grouped = build_monthly_trend_narrative(
        raw_analysis, fx_analysis, compound_df
    )
    st.markdown(
        '<div class="trend-report">'
        '<div class="trend-report-title">월간 원자재 동향 종합</div>'
        f'<div class="trend-report-body">{html.escape(narrative)}</div>'
        '</div>',
        unsafe_allow_html=True,
    )

    rising = grouped[grouped["추세 구분"].isin(["지속 상승", "하락 후 반등"])]
    falling = grouped[grouped["추세 구분"].isin(["지속 하락", "상승 후 조정"])]
    turning = grouped[grouped["추세 구분"].isin(["상승 후 조정", "하락 후 반등"])]
    stable = grouped[grouped["추세 구분"] == "보합·혼조"]
    columns = st.columns(4)
    with columns[0]:
        render_trend_group_card("상승·반등", rising, "red", "지속 상승 또는 하락 후 반등")
    with columns[1]:
        render_trend_group_card("하락·조정", falling, "blue", "지속 하락 또는 상승 후 조정")
    with columns[2]:
        render_trend_group_card("방향 전환", turning, "orange", "단기와 중기 방향이 달라진 품목")
    with columns[3]:
        render_trend_group_card("보합·혼조", stable, "gray", "뚜렷한 방향성이 제한된 품목")

    st.markdown("### 2. 주요 품목 전월 동향")
    render_monthly_change_chart(raw_analysis)

    st.markdown("### 3. 품목별 추세 변화")
    trend_table = grouped[
        ["품목", "관련 구매품", "전월 증감률", "6개월 증감률"]
    ].copy()
    trend_table["추세 구분"] = grouped["추세 구분"]
    trend_order = {
        "지속 상승": 0,
        "하락 후 반등": 1,
        "상승 후 조정": 2,
        "지속 하락": 3,
        "보합·혼조": 4,
        "데이터 부족": 5,
    }
    trend_table["_order"] = trend_table["추세 구분"].map(trend_order)
    trend_table = trend_table.sort_values(["_order", "전월 증감률"], ascending=[True, False])
    st.dataframe(
        trend_table.drop(columns="_order"),
        hide_index=True,
        use_container_width=True,
        column_config={
            "전월 증감률": st.column_config.NumberColumn(format="%+.2f%%"),
            "6개월 증감률": st.column_config.NumberColumn(format="%+.2f%%"),
            "추세 구분": st.column_config.TextColumn(
                help="전월과 6개월 증감률의 방향을 조합한 현재 추세"
            ),
        },
    )
    st.caption(
        "지속 상승·지속 하락은 전월과 6개월 방향이 같은 경우이며, "
        "상승 후 조정·하락 후 반등은 단기 방향이 중기 흐름과 달라진 경우입니다."
    )

    st.markdown("### 4. 환율 동향")
    if fx_analysis.empty:
        st.info("환율 데이터가 없습니다.")
    else:
        render_monthly_change_chart(fx_analysis)

    if not compound_df.empty:
        with st.expander("원자재와 결제 통화의 동반 움직임 보기"):
            st.dataframe(
                compound_df,
                hide_index=True,
                use_container_width=True,
                column_config={
                    "원자재 전월 증감률": st.column_config.NumberColumn(format="%+.2f%%"),
                    "해당 환율 전월 증감률": st.column_config.NumberColumn(format="%+.2f%%"),
                },
            )

    with st.expander("상세 동향 지표 보기"):
        detail = analysis[
            [
                "구분", "품목", "관련 구매품", "최근 기준월", "전월 증감률",
                "3개월 평균 대비", "6개월 증감률", "전년 평균 대비",
                "12개월 범위 위치", "6개월 변동성", "데이터 상태",
            ]
        ].copy()
        detail["최근 기준월"] = detail["최근 기준월"].map(
            lambda value: f"{value.year}년 {value.month}월"
        )
        st.dataframe(
            detail,
            hide_index=True,
            use_container_width=True,
            column_config={
                "전월 증감률": st.column_config.NumberColumn(format="%+.2f%%"),
                "3개월 평균 대비": st.column_config.NumberColumn(format="%+.2f%%"),
                "6개월 증감률": st.column_config.NumberColumn(format="%+.2f%%"),
                "전년 평균 대비": st.column_config.NumberColumn(format="%+.2f%%"),
                "12개월 범위 위치": st.column_config.ProgressColumn(
                    min_value=0, max_value=100, format="%.0f%%"
                ),
                "6개월 변동성": st.column_config.NumberColumn(format="%.2f%%"),
            },
        )

    if len(stale_items):
        stale_names = ", ".join(stale_items["품목"].astype(str))
        st.warning(f"최신 기준월이 다른 항목: {stale_names}")
    st.caption(
        "※ 품목마다 단위와 사용 방향이 다르므로 전체 원자재를 하나의 평균 수치로 합산하지 않습니다. "
        "미래 가격 예측이나 외부 시장 전망은 포함하지 않습니다."
    )


def trend_note(label):
    notes = {
        "지속 상승": "전월과 최근 6개월이 모두 상승 방향입니다.",
        "지속 하락": "전월과 최근 6개월이 모두 하락 방향입니다.",
        "상승 후 조정": "중기 상승 흐름에서 최근 가격이 조정됐습니다.",
        "하락 후 반등": "중기 하락 흐름에서 최근 가격이 반등했습니다.",
        "보합·혼조": "단기와 중기 방향성이 뚜렷하지 않습니다.",
        "데이터 부족": "추세 판단에 필요한 월별 데이터가 부족합니다.",
    }
    return notes.get(label, "")


def build_briefing_lines(grouped, compound_df):
    lines = []
    valid = grouped.dropna(subset=["전월 증감률"])
    rising = valid[valid["전월 증감률"] > 0]
    falling = valid[valid["전월 증감률"] < 0]
    if not rising.empty:
        row = rising.loc[rising["전월 증감률"].idxmax()]
        lines.append(f"가장 큰 상승: {row['품목']} 전월 대비 {row['전월 증감률']:+.1f}%")
    if not falling.empty:
        row = falling.loc[falling["전월 증감률"].idxmin()]
        lines.append(f"가장 큰 하락: {row['품목']} 전월 대비 {row['전월 증감률']:+.1f}%")

    turning = grouped[grouped["추세 구분"].isin(["상승 후 조정", "하락 후 반등"])]
    if not turning.empty:
        details = ", ".join(
            f"{row['품목']}({row['추세 구분']})" for _, row in turning.iterrows()
        )
        lines.append(f"방향이 달라진 품목: {details}")
    else:
        lines.append("방향 전환: 전월과 중기 방향이 엇갈린 주요 품목 없음")

    if not compound_df.empty:
        amplified = compound_df[
            compound_df["동반 영향"].isin(
                ["원자재·환율 동반 상승", "환율 상승이 원자재 하락 효과 일부 축소"]
            )
        ]
        offset = compound_df[compound_df["동반 영향"] == "환율 하락이 원자재 상승 일부 상쇄"]
        if not amplified.empty:
            lines.append(
                "환율 영향: " + ", ".join(amplified["원자재"].astype(str))
                + "의 원화 환산 부담을 확대하는 방향"
            )
        elif not offset.empty:
            lines.append(
                "환율 영향: " + ", ".join(offset["원자재"].astype(str))
                + "의 가격 상승을 일부 상쇄하는 방향"
            )
        else:
            lines.append("환율 영향: 원자재 가격과 환율의 뚜렷한 동반 상승 없음")
    return lines


def render_monthly_highlights(grouped):
    valid = grouped.dropna(subset=["전월 증감률"])
    biggest_rise = valid[valid["전월 증감률"] > 0].sort_values(
        "전월 증감률", ascending=False
    ).head(1)
    biggest_fall = valid[valid["전월 증감률"] < 0].sort_values(
        "전월 증감률"
    ).head(1)
    smallest_change = (
        valid.assign(_absolute_change=valid["전월 증감률"].abs())
        .sort_values("_absolute_change")
        .head(1)
    )

    columns = st.columns(3)
    cards = [
        ("가장 큰 상승", biggest_rise, "red"),
        ("가장 큰 하락", biggest_fall, "blue"),
        ("변동이 가장 작은 품목", smallest_change, "orange"),
    ]
    for column, (title, frame, color) in zip(columns, cards):
        with column:
            if frame.empty:
                main = "해당 없음"
                note = "이번 달 해당 품목이 없습니다."
            else:
                row = frame.iloc[0]
                main = str(row["품목"])
                if abs(row["전월 증감률"]) < 0.01:
                    note = "전월 대비 ― 0.0%"
                else:
                    arrow = "▲" if row["전월 증감률"] > 0 else "▼"
                    note = f"전월 대비 {arrow} {abs(row['전월 증감률']):.1f}%"
            st.markdown(
                f'<div class="trend-group-card {color}">'
                f'<div class="trend-group-title">{html.escape(title)}</div>'
                f'<div class="trend-group-items">{html.escape(main)}</div>'
                f'<div class="trend-group-note">{html.escape(note)}</div>'
                '</div>',
                unsafe_allow_html=True,
            )


def render_item_briefing_cards(grouped, show_related=True, show_trend=True):
    cards = []
    for _, row in grouped.sort_values("전월 증감률", ascending=False, na_position="last").iterrows():
        mom = row["전월 증감률"]
        if pd.isna(mom) or abs(mom) < 0.01:
            color = "flat"
            change_text = "전월 대비 ― 0.0%" if pd.notna(mom) else "전월 데이터 없음"
        elif mom > 0:
            color = "up"
            change_text = f"전월 대비 ▲ {abs(mom):.1f}%"
        else:
            color = "down"
            change_text = f"전월 대비 ▼ {abs(mom):.1f}%"
        value_unit = combined_unit(row["통화"], row["단위"])
        note = str(row.get("비고", "") or "").strip()
        note_html = (
            f'<div class="briefing-card-note">{html.escape(note)}</div>'
            if note else ""
        )
        related_html = (
            f'<div class="briefing-card-related">{html.escape(str(row["관련 구매품"]))}</div>'
            if show_related else ""
        )
        trend_html = ""
        if show_trend:
            trend_html = (
                f'<div class="briefing-card-trend">{html.escape(str(row["추세 구분"]))}</div>'
                f'<div class="briefing-card-note">{html.escape(trend_note(row["추세 구분"]))}</div>'
            )
        cards.append(
            f'<div class="briefing-card {color}">'
            f'<div class="briefing-card-name">{html.escape(str(row["품목"]))}</div>'
            f'{note_html}'
            f'{related_html}'
            f'<span class="briefing-card-value">{html.escape(format_value(row["최신값"], value_unit))}</span>'
            f'<span class="briefing-card-change {color}">{html.escape(change_text)}</span>'
            f'{trend_html}'
            '</div>'
        )
    st.markdown(
        f'<div class="briefing-grid">{"".join(cards)}</div>',
        unsafe_allow_html=True,
    )


def render_monthly_briefing(datasets):
    analysis = build_procurement_metrics(datasets)
    if analysis.empty:
        st.info("월간 동향을 표시할 데이터가 없습니다.")
        return

    raw = analysis[analysis["구분"] == "원자재"].copy()
    fx = analysis[analysis["구분"] == "환율"].copy()
    latest_date = raw["최근 기준월"].max()
    valid = raw.dropna(subset=["전월 증감률"]).copy()
    rising = valid[valid["전월 증감률"] > 0.01].sort_values(
        "전월 증감률", ascending=False
    )
    falling = valid[valid["전월 증감률"] < -0.01].sort_values("전월 증감률")
    flat = valid[valid["전월 증감률"].abs() <= 0.01]

    if not rising.empty and not falling.empty:
        top_rise = rising.iloc[0]
        top_fall = falling.iloc[0]
        overall = (
            f"{top_rise['품목']}가 전월 대비 {top_rise['전월 증감률']:+.1f}%로 가장 크게 상승했고, "
            f"{top_fall['품목']}는 {top_fall['전월 증감률']:+.1f}%로 가장 크게 하락했습니다. "
            "주요 원자재의 움직임은 품목별로 엇갈렸습니다."
        )
    elif not rising.empty:
        top_rise = rising.iloc[0]
        overall = (
            f"{top_rise['품목']}가 전월 대비 {top_rise['전월 증감률']:+.1f}%로 "
            "가장 큰 상승폭을 기록했습니다."
        )
    elif not falling.empty:
        top_fall = falling.iloc[0]
        overall = (
            f"{top_fall['품목']}가 전월 대비 {top_fall['전월 증감률']:+.1f}%로 "
            "가장 큰 하락폭을 기록했습니다."
        )
    else:
        overall = "확인 가능한 원자재 가격은 전월과 비슷한 수준을 유지했습니다."

    def report_item(row, show_related=True):
        change = row["전월 증감률"]
        color = "#d32f2f" if change > 0.01 else "#1976d2" if change < -0.01 else "#616161"
        card_class = "up" if change > 0.01 else "down" if change < -0.01 else "flat"
        arrow = "▲" if change > 0.01 else "▼" if change < -0.01 else "―"
        related = str(row.get("관련 구매품", "") or "-")
        value_unit = combined_unit(row.get("통화"), row.get("단위"))
        three_month = row.get("3개월 평균 대비")
        six_month = row.get("6개월 증감률")

        def comparison_html(label, value):
            if pd.isna(value):
                return f'<span style="color:#9ca3af">{label} 비교 불가</span>'
            compare_color = "#d32f2f" if value > 0.01 else "#1976d2" if value < -0.01 else "#616161"
            compare_arrow = "▲" if value > 0.01 else "▼" if value < -0.01 else "―"
            return (
                f'<span style="color:{compare_color};font-weight:750">'
                f'{label} {compare_arrow} {abs(value):.1f}%</span>'
            )

        related_html = (
            f'<div class="briefing-card-related">관련 구매품 · {html.escape(related)}</div>'
            if show_related else ""
        )
        return (
            f'<div class="briefing-card {card_class}">'
            f'<div class="briefing-card-name">{html.escape(str(row["품목"]))}</div>'
            f'{related_html}'
            f'<div class="briefing-card-value" style="display:block;margin-top:14px">'
            f'{html.escape(format_value(row["최신값"], value_unit))}</div>'
            f'<div style="margin-top:8px;color:{color};font-weight:850">'
            f'전월 대비 {arrow} {abs(change):.1f}%</div>'
            f'<div style="display:flex;flex-wrap:wrap;gap:6px 12px;margin-top:10px;'
            f'padding-top:9px;border-top:1px solid #edf0f4;font-size:.82rem">'
            f'{comparison_html("3개월 평균 대비", three_month)}'
            f'{comparison_html("6개월 대비", six_month)}'
            '</div></div>'
        )

    key_items = valid.sort_values("품목")
    key_items_html = "".join(report_item(row) for _, row in key_items.iterrows())

    fx_valid = fx.dropna(subset=["전월 증감률"]).copy()
    if fx_valid.empty:
        fx_sentence = "환율은 비교 가능한 전월 데이터가 없습니다."
    else:
        fx_parts = []
        for _, row in fx_valid.sort_values("품목").iterrows():
            change = row["전월 증감률"]
            arrow = "▲" if change > 0.01 else "▼" if change < -0.01 else "―"
            fx_parts.append(f"{row['품목']} {arrow}{abs(change):.1f}%")
        fx_sentence = "전월 대비 " + ", ".join(fx_parts) + " 변동했습니다."

    medium_valid = raw.dropna(subset=["6개월 증감률"]).copy()
    if medium_valid.empty:
        medium_sentence = "6개월 흐름은 비교 가능한 데이터가 없습니다."
    else:
        medium_row = medium_valid.loc[medium_valid["6개월 증감률"].abs().idxmax()]
        medium_direction = "상승" if medium_row["6개월 증감률"] > 0 else "하락" if medium_row["6개월 증감률"] < 0 else "보합"
        medium_sentence = (
            f"6개월 기준 변동폭이 가장 큰 품목은 {medium_row['품목']}으로 "
            f"{abs(medium_row['6개월 증감률']):.1f}% {medium_direction}했습니다."
        )

    related_parts = []
    if not rising.empty:
        row = rising.iloc[0]
        related_parts.append(f"{row['품목']} → {row['관련 구매품']}")
    if not falling.empty:
        row = falling.iloc[0]
        related_parts.append(f"{row['품목']} → {row['관련 구매품']}")
    related_sentence = (
        "주요 변동 원자재의 관련 구매품은 " + "; ".join(related_parts) + "입니다."
        if related_parts else "주요 변동 원자재와 연결된 관련 구매품이 없습니다."
    )

    def colored_change(value):
        if pd.isna(value) or abs(value) <= 0.01:
            return '<strong style="color:#616161">― 0.0%</strong>'
        change_color = "#d32f2f" if value > 0 else "#1976d2"
        arrow = "▲" if value > 0 else "▼"
        return f'<strong style="color:{change_color}">{arrow} {abs(value):.1f}%</strong>'

    if not rising.empty and not falling.empty:
        rise_row = rising.iloc[0]
        fall_row = falling.iloc[0]
        overall_html = (
            f'{html.escape(str(rise_row["품목"]))}가 전월 대비 '
            f'{colored_change(rise_row["전월 증감률"])}로 가장 크게 상승했고, '
            f'{html.escape(str(fall_row["품목"]))}는 '
            f'{colored_change(fall_row["전월 증감률"])}로 가장 크게 하락했습니다.'
        )
    elif not rising.empty:
        rise_row = rising.iloc[0]
        overall_html = (
            f'{html.escape(str(rise_row["품목"]))}가 전월 대비 '
            f'{colored_change(rise_row["전월 증감률"])}로 가장 큰 상승폭을 기록했습니다.'
        )
    elif not falling.empty:
        fall_row = falling.iloc[0]
        overall_html = (
            f'{html.escape(str(fall_row["품목"]))}가 전월 대비 '
            f'{colored_change(fall_row["전월 증감률"])}로 가장 큰 하락폭을 기록했습니다.'
        )
    else:
        overall_html = html.escape(overall)

    if medium_valid.empty:
        medium_html = html.escape(medium_sentence)
    else:
        medium_html = (
            f'6개월 기준 변동폭이 가장 큰 품목은 {html.escape(str(medium_row["품목"]))}으로 '
            f'{colored_change(medium_row["6개월 증감률"])} 변동했습니다.'
        )

    if fx_valid.empty:
        fx_html = html.escape(fx_sentence)
    else:
        fx_segments = []
        for _, fx_row in fx_valid.sort_values("품목").iterrows():
            fx_segments.append(
                f'{html.escape(str(fx_row["품목"]))} {colored_change(fx_row["전월 증감률"])}'
            )
        fx_html = "전월 대비 " + ", ".join(fx_segments) + " 변동했습니다."

    year_valid = raw.dropna(subset=["전년 평균 대비"]).copy()
    if year_valid.empty:
        year_html = "전년 평균과 비교할 수 있는 데이터가 없습니다."
    else:
        year_row = year_valid.loc[year_valid["전년 평균 대비"].abs().idxmax()]
        year_html = (
            f'전년 평균 대비 변동폭이 가장 큰 품목은 {html.escape(str(year_row["품목"]))}으로 '
            f'{colored_change(year_row["전년 평균 대비"])} 차이를 보였습니다.'
        )

    def summary_line(label, content_html, color, background):
        return (
            '<div style="display:grid;grid-template-columns:105px 1fr;gap:12px;align-items:start;'
            'padding:10px 0;border-bottom:1px solid #e9edf2">'
            f'<span style="display:inline-block;width:max-content;padding:4px 9px;border-radius:7px;'
            f'background:{background};color:{color};font-weight:850;font-size:.84rem">'
            f'{html.escape(label)}</span>'
            f'<span style="color:#374151;line-height:1.65">{content_html}</span></div>'
        )

    summary_html = (
        summary_line("전월 동향", overall_html, "#172033", "#e9eef5")
        + summary_line("중기 흐름", medium_html, "#6d4c41", "#f5eee9")
        + summary_line("전년 비교", year_html, "#5d4037", "#f4ece8")
        + summary_line("관련 구매품", html.escape(related_sentence), "#455a64", "#eef3f5")
        + summary_line("환율", fx_html, "#37474f", "#edf2f4")
    )

    st.subheader(f"{latest_date.year}년 {latest_date.month}월 시세 동향")
    st.caption("원자재·환율 시세와 관련 구매품을 연결해 구매팀 참고사항을 자동 분석합니다.")
    st.markdown(
        '<div style="border:1px solid #dfe4ea;border-radius:14px;background:#fff;overflow:hidden;'
        'box-shadow:0 2px 8px rgba(15,23,42,.04);margin:10px 0 24px">'
        '<div style="padding:18px 20px;background:#f7f9fc;border-bottom:1px solid #e5e9ef">'
        '<div style="font-size:1.18rem;font-weight:850;color:#172033">핵심 요약</div>'
        f'<div style="margin-top:7px">{summary_html}</div>'
        '</div>'
        '<div style="padding:16px 20px 4px;font-size:1rem;font-weight:800;color:#172033">'
        '품목별 원자재 동향</div>'
        f'<div class="briefing-grid" style="padding:12px 20px 18px">{key_items_html}</div>'
        '<div style="padding:12px 20px 16px;color:#6b7280;font-size:.84rem">'
        '※ 관련 구매품은 해당 원자재 가격 변동의 참고 대상이며, 실제 구매가격 변동률을 의미하지 않습니다.'
        '</div></div>',
        unsafe_allow_html=True,
    )

    stale = analysis[analysis["데이터 상태"] == "업데이트 필요"]
    if not stale.empty:
        st.warning(
            f"기준월 확인: {', '.join(stale['품목'].astype(str))}은 다른 항목보다 "
            "최신 기준월이 이전입니다."
        )

    st.markdown("### 원자재 전월 변화")
    render_monthly_change_chart(raw)

    if not fx.empty:
        st.divider()
        fx_latest = fx["최근 기준월"].max()
        st.markdown(f"### {fx_latest.year}년 {fx_latest.month}월 환율 동향")
        st.caption("각 통화의 최신 환율과 전월·3개월 평균·6개월 변화를 표시합니다.")
        fx_items_html = "".join(
            report_item(row, show_related=False)
            for _, row in fx_valid.sort_values("품목").iterrows()
        )
        st.markdown(
            f'<div class="briefing-grid" style="margin:10px 0 20px">{fx_items_html}</div>',
            unsafe_allow_html=True,
        )
        render_monthly_change_chart(fx)

    st.caption(
        "※ 업로드된 가격을 바탕으로 한 참고용 동향입니다. 실제 구매가격과 원자재·환율의 영향 정도는 "
        "계약조건, 적용 시점 및 원가 구성에 따라 달라질 수 있습니다."
    )


def render_beta_analysis(datasets):
    analysis = build_procurement_metrics(datasets)
    if analysis.empty:
        st.info("자동 분석에 사용할 데이터가 없습니다.")
        return
    raw = analysis[analysis["구분"] == "원자재"].copy()
    fx = analysis[analysis["구분"] == "환율"].copy()
    history_all = datasets.get("원자재", pd.DataFrame())
    if raw.empty:
        st.info("원자재 데이터가 없어 자동 분석을 표시할 수 없습니다.")
        return

    def pct(value):
        if pd.isna(value):
            return "자료 없음"
        arrow = "▲" if value > 0 else "▼" if value < 0 else "―"
        return f"{arrow} {abs(float(value)):.1f}%"

    def tone(value):
        if pd.isna(value) or abs(float(value)) < .01:
            return "#667085"
        return "#dc2626" if float(value) > 0 else "#1677d2"

    def names(frame):
        return ", ".join(frame["품목"].astype(str).tolist()) if not frame.empty else "없음"

    st.subheader("분석 인사이트")
    st.caption("BETA · 업로드된 가격·환율의 실제 변화만 분석합니다. 외부 전망이나 구매단가를 추정하지 않습니다.")
    beta_tabs = st.tabs(["종합 의견", "흐름 분석", "12개월 가격 추이", "현재 가격 위치"])

    with beta_tabs[0]:
        comparable = raw.dropna(subset=["전월 증감률", "6개월 증감률"]).copy()
        continued_up = comparable[(comparable["전월 증감률"] > 0) & (comparable["6개월 증감률"] > 0)]
        continued_down = comparable[(comparable["전월 증감률"] < 0) & (comparable["6개월 증감률"] < 0)]
        correction = comparable[(comparable["전월 증감률"] < 0) & (comparable["6개월 증감률"] > 0)]
        rebound = comparable[(comparable["전월 증감률"] > 0) & (comparable["6개월 증감률"] < 0)]
        mom_valid = raw.dropna(subset=["전월 증감률"])
        six_valid = raw.dropna(subset=["6개월 증감률"])
        rise = mom_valid.sort_values("전월 증감률", ascending=False).head(1)
        fall = mom_valid.sort_values("전월 증감률").head(1)
        volatile = raw.sort_values("6개월 변동성", ascending=False).head(1)
        high_band = raw.sort_values("12개월 범위 위치", ascending=False).head(1)
        low_band = raw.sort_values("12개월 범위 위치").head(1)

        rise_text = (f'{rise.iloc[0]["품목"]} {pct(rise.iloc[0]["전월 증감률"])}' if not rise.empty else "자료 없음")
        fall_text = (f'{fall.iloc[0]["품목"]} {pct(fall.iloc[0]["전월 증감률"])}' if not fall.empty else "자료 없음")
        if not mom_valid.empty and (mom_valid["전월 증감률"] > 0).all():
            market_shape = "확인 가능한 원자재가 모두 전월보다 상승해 단기적으로 상승 방향이 뚜렷합니다."
        elif not mom_valid.empty and (mom_valid["전월 증감률"] < 0).all():
            market_shape = "확인 가능한 원자재가 모두 전월보다 하락해 단기적으로 하락 방향이 뚜렷합니다."
        else:
            market_shape = "원자재별 상승과 하락이 엇갈려 시장 전체가 한 방향으로 움직이기보다 품목별 차별화가 나타났습니다."

        momentum_parts = []
        if not continued_up.empty:
            momentum_parts.append(f"{names(continued_up)}은 상승 흐름이 이어지고 있습니다")
        if not correction.empty:
            momentum_parts.append(f"{names(correction)}은 중기 상승 후 최근 조정을 보였습니다")
        if not rebound.empty:
            momentum_parts.append(f"{names(rebound)}은 중기 하락 후 최근 반등했습니다")
        if not continued_down.empty:
            momentum_parts.append(f"{names(continued_down)}은 하락 흐름이 이어지고 있습니다")
        momentum_summary = ". ".join(momentum_parts) + ("." if momentum_parts else "기간별 흐름을 비교할 데이터가 부족합니다.")

        if not high_band.empty and not low_band.empty and not volatile.empty:
            hi, lo, vol = high_band.iloc[0], low_band.iloc[0], volatile.iloc[0]
            range_summary = (
                f'{hi["품목"]}은 최근 12개월 가격 범위의 상단({hi["12개월 범위 위치"]:.0f}%)에 있고, '
                f'{lo["품목"]}은 하단({lo["12개월 범위 위치"]:.0f}%)에 있습니다. '
                f'월별 움직임은 {vol["품목"]}의 변동성({vol["6개월 변동성"]:.1f}%)이 가장 컸습니다.'
            )
        else:
            range_summary = "가격 위치와 변동성을 함께 판단할 데이터가 부족합니다."

        rise_html = f'<b style="color:#dc2626;">{html.escape(rise_text)}</b>'
        fall_html = f'<b style="color:#1677d2;">{html.escape(fall_text)}</b>'
        overall_html = (
            f'{html.escape(market_shape)} 가장 큰 단기 움직임은 '
            f'<span style="color:#dc2626;font-weight:850;">상승 {rise_html}</span>, '
            f'<span style="color:#1677d2;font-weight:850;">하락 {fall_html}</span>입니다. '
            '전체 평균보다 품목별 상승 지속·최근 하락·최근 상승 여부를 구분해 보는 것이 핵심입니다.'
        )
        momentum_bits = []
        if not continued_up.empty:
            momentum_bits.append(f'<span style="color:#dc2626;font-weight:850;">상승 지속 · {html.escape(names(continued_up))}</span>')
        if not correction.empty:
            momentum_bits.append(f'<span style="color:#9a5b13;font-weight:850;">최근 하락(중기 상승) · {html.escape(names(correction))}</span>')
        if not rebound.empty:
            momentum_bits.append(f'<span style="color:#7c3aed;font-weight:850;">최근 상승(중기 하락) · {html.escape(names(rebound))}</span>')
        if not continued_down.empty:
            momentum_bits.append(f'<span style="color:#1677d2;font-weight:850;">하락 지속 · {html.escape(names(continued_down))}</span>')
        momentum_html = " &nbsp;│&nbsp; ".join(momentum_bits) if momentum_bits else html.escape(momentum_summary)

        if not high_band.empty and not low_band.empty and not volatile.empty:
            hi, lo, vol = high_band.iloc[0], low_band.iloc[0], volatile.iloc[0]
            range_html = (
                f'<span style="color:#dc2626;font-weight:850;">상단 · {html.escape(str(hi["품목"]))} '
                f'{hi["12개월 범위 위치"]:.0f}%</span> &nbsp;│&nbsp; '
                f'<span style="color:#1677d2;font-weight:850;">하단 · {html.escape(str(lo["품목"]))} '
                f'{lo["12개월 범위 위치"]:.0f}%</span> &nbsp;│&nbsp; '
                f'<span style="color:#7c3aed;font-weight:850;">변동성 최대 · {html.escape(str(vol["품목"]))} '
                f'{vol["6개월 변동성"]:.1f}%</span>'
            )
        else:
            range_html = html.escape(range_summary)

        purchase_bits = []
        if not rise.empty:
            row = rise.iloc[0]
            purchase_bits.append(
                f'<span style="color:#dc2626;font-weight:850;">{html.escape(str(row["품목"]))} ▲</span> '
                f'→ {html.escape(str(row["관련 구매품"]))}'
            )
        if not fall.empty:
            row = fall.iloc[0]
            purchase_bits.append(
                f'<span style="color:#1677d2;font-weight:850;">{html.escape(str(row["품목"]))} ▼</span> '
                f'→ {html.escape(str(row["관련 구매품"]))}'
            )
        purchase_html = " &nbsp;│&nbsp; ".join(purchase_bits) if purchase_bits else "관련 구매품 연결 정보가 부족합니다."
        fx_bits = []
        for _, row in fx.iterrows():
            value = row["전월 증감률"]
            fx_bits.append(
                f'<span style="color:{tone(value)};font-weight:850;">'
                f'{html.escape(str(row["품목"]))} {html.escape(pct(value))}</span>'
            )
        fx_html = " &nbsp;│&nbsp; ".join(fx_bits) if fx_bits else "업로드된 환율 데이터가 없습니다."
        summary_rows = [
            ("전체 흐름", overall_html, "#344054"),
            ("기간별 방향", momentum_html, "#9a5b13"),
            ("가격 위치·변동", range_html, "#7c3aed"),
            ("관련 구매품", purchase_html, "#0f766e"),
            ("환율 흐름", fx_html, "#475467"),
        ]
        st.markdown("#### 한눈에 보는 전체 요약")
        st.markdown(
            '<div style="border:1px solid #dfe5ec;border-radius:16px;overflow:hidden;background:white;'
            'box-shadow:0 4px 16px rgba(16,24,40,.05);margin-bottom:16px;">'
            '<div style="padding:16px 18px;background:linear-gradient(135deg,#f4f7fb,#fff);border-bottom:1px solid #e4e7ec;">'
            '<div style="font-size:12px;color:#667085;font-weight:800;">업로드 수치 종합 판독</div>'
            f'<div style="font-size:19px;font-weight:900;color:#101828;margin-top:5px;">{html.escape(market_shape)}</div></div>'
            + "".join(
                f'<div style="display:grid;grid-template-columns:118px 1fr;gap:14px;padding:13px 17px;'
                f'border-bottom:1px solid #eef1f5;align-items:start;">'
                f'<div style="font-weight:850;color:{accent};">{title}</div>'
                f'<div style="line-height:1.7;color:#344054;">{body}</div></div>'
                for title, body, accent in summary_rows
            ) + '</div>', unsafe_allow_html=True,
        )

        flow_cards = [
            ("상승 흐름 지속", names(continued_up), "전월과 6개월 흐름이 모두 상승", "#dc2626"),
            ("최근 하락(중기 상승)", names(correction), "6개월 상승 후 최근월 하락", "#9a5b13"),
            ("최근 상승(중기 하락)", names(rebound), "6개월 하락 후 최근월 상승", "#7c3aed"),
            ("하락 흐름 지속", names(continued_down), "전월과 6개월 흐름이 모두 하락", "#1677d2"),
        ]
        st.markdown("#### 흐름 구조")
        st.markdown(
            '<div style="display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:10px;">' +
            "".join(
                f'<div style="border:1px solid #e4e7ec;border-top:4px solid {accent};border-radius:13px;padding:14px;background:white;">'
                f'<div style="font-size:12px;color:{accent};font-weight:850;">{title}</div>'
                f'<div style="font-size:17px;font-weight:900;margin:9px 0;color:#101828;min-height:48px;">{html.escape(item_names)}</div>'
                f'<div style="font-size:11px;color:#667085;line-height:1.5;">{description}</div></div>'
                for title, item_names, description, accent in flow_cards
            ) + '</div>', unsafe_allow_html=True,
        )

        st.markdown("#### 핵심 관찰")
        observations = []
        if not six_valid.empty:
            long_move = six_valid.loc[six_valid["6개월 증감률"].abs().idxmax()]
            observations.append(("중기 변화", f'{long_move["품목"]}의 6개월 변화가 {pct(long_move["6개월 증감률"])}로 가장 컸습니다.', tone(long_move["6개월 증감률"])))
        if not volatile.empty:
            row = volatile.iloc[0]
            observations.append(("변동성", f'{row["품목"]}의 최근 6개월 변동성은 {row["6개월 변동성"]:.1f}%로 품목 중 가장 높습니다.', "#7c3aed"))
        if not high_band.empty and not low_band.empty:
            hi, lo = high_band.iloc[0], low_band.iloc[0]
            observations.append(("가격 위치", f'{hi["품목"]}은 12개월 범위의 {hi["12개월 범위 위치"]:.0f}% 지점, {lo["품목"]}은 {lo["12개월 범위 위치"]:.0f}% 지점에 있습니다.', "#0f766e"))
        stale = raw[raw["데이터 상태"] != "최신"]
        observations.append(("데이터 상태", f'최근 기준월보다 업데이트가 늦은 품목: {names(stale)}.', "#667085"))
        st.markdown(
            '<div style="border:1px solid #e4e7ec;border-radius:14px;overflow:hidden;background:white;">' +
            "".join(
                f'<div style="display:grid;grid-template-columns:120px 1fr;gap:12px;padding:13px 16px;border-bottom:1px solid #eef1f5;">'
                f'<b style="color:{accent};">{title}</b><span style="color:#344054;line-height:1.6;">{html.escape(body)}</span></div>'
                for title, body, accent in observations
            ) + '</div>', unsafe_allow_html=True,
        )

        st.markdown("#### 구매품·환율 관점")
        changed = raw.dropna(subset=["전월 증감률"]).copy()
        changed = changed.reindex(changed["전월 증감률"].abs().sort_values(ascending=False).index).head(3)
        purchase_text = " / ".join(
            f'{row["품목"]} → {row["관련 구매품"]} ({pct(row["전월 증감률"])})'
            for _, row in changed.iterrows()
        ) or "관련 구매품을 연결할 데이터가 없습니다."
        fx_text = " / ".join(
            f'{row["품목"]} {pct(row["전월 증감률"])}' for _, row in fx.iterrows()
        ) or "환율 데이터가 없습니다."
        st.markdown(
            f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;">'
            f'<div style="padding:16px;border-radius:13px;background:#f8fafc;border:1px solid #e4e7ec;">'
            f'<b>관련 구매품 참고 신호</b><div style="margin-top:8px;line-height:1.65;color:#475467;">{html.escape(purchase_text)}</div></div>'
            f'<div style="padding:16px;border-radius:13px;background:#f8fafc;border:1px solid #e4e7ec;">'
            f'<b>결제 통화 흐름</b><div style="margin-top:8px;line-height:1.65;color:#475467;">{html.escape(fx_text)}</div></div></div>',
            unsafe_allow_html=True,
        )
        st.info("관련 구매품은 원자재 가격 변화의 참고 대상입니다. 실제 구매단가 반영 여부·반영률·시차는 계약조건과 원가구조를 별도로 확인해야 합니다.")

    with beta_tabs[1]:
        st.markdown("#### 품목별 가격 흐름")
        st.caption("최신 가격과 단기·중기 방향을 함께 읽습니다. 방향은 가격의 유불리가 아니라 움직임을 의미합니다.")
        st.markdown(
            '<div style="display:flex;flex-wrap:wrap;gap:7px;margin:3px 0 14px;font-size:11px;">'
            '<span title="전월 가격과 6개월 전 가격보다 모두 상승" style="padding:6px 9px;border-radius:8px;background:#fff1f1;color:#dc2626;"><b>상승 지속</b> · 전월↑ / 6개월↑</span>'
            '<span title="6개월 전보다 높지만 전월보다 하락" style="padding:6px 9px;border-radius:8px;background:#fff7ed;color:#9a5b13;"><b>최근 하락</b> · 전월↓ / 6개월↑</span>'
            '<span title="6개월 전보다 낮지만 전월보다 상승" style="padding:6px 9px;border-radius:8px;background:#f5f3ff;color:#7c3aed;"><b>최근 상승</b> · 전월↑ / 6개월↓</span>'
            '<span title="전월 가격과 6개월 전 가격보다 모두 하락" style="padding:6px 9px;border-radius:8px;background:#eef6ff;color:#1677d2;"><b>하락 지속</b> · 전월↓ / 6개월↓</span>'
            '<span style="padding:6px 2px;color:#98a2b3;">※ 배지에 마우스를 올리면 설명이 표시됩니다.</span></div>',
            unsafe_allow_html=True,
        )
        flow_cards = []
        for _, row in raw.iterrows():
            mom = row["전월 증감률"]
            mid = row["3개월 평균 대비"]
            six = row["6개월 증감률"]
            if pd.notna(mom) and pd.notna(six):
                if mom > 0 and six > 0:
                    status, status_color, interpretation = "상승 흐름 지속", "#dc2626", "단기와 중기 흐름이 모두 상승 방향입니다."
                elif mom < 0 and six < 0:
                    status, status_color, interpretation = "하락 흐름 지속", "#1677d2", "단기와 중기 흐름이 모두 하락 방향입니다."
                elif mom < 0 and six > 0:
                    status, status_color, interpretation = "최근 하락(중기 상승)", "#9a5b13", "6개월 전보다 높지만 전월보다 낮아 최근에는 하락 방향입니다."
                elif mom > 0 and six < 0:
                    status, status_color, interpretation = "최근 상승(중기 하락)", "#7c3aed", "6개월 전보다 낮지만 전월보다 높아 최근에는 상승 방향입니다."
                else:
                    status, status_color, interpretation = "보합 구간", "#667085", "최근 가격 방향이 뚜렷하지 않습니다."
            else:
                status, status_color, interpretation = "비교 자료 부족", "#667085", "기간별 흐름을 판단할 데이터가 부족합니다."
            display_unit = f'{row["통화"]} / {row["단위"]}'
            metrics = [
                ("전월", mom), ("3개월 평균", mid), ("6개월", six), ("전년 평균", row["전년 평균 대비"]),
            ]
            metric_html = "".join(
                f'<div style="padding:9px 8px;border-radius:9px;background:#f8fafc;text-align:center;">'
                f'<div style="font-size:10px;color:#667085;margin-bottom:4px;">{label}</div>'
                f'<div style="font-size:14px;font-weight:850;color:{tone(value)};">{pct(value)}</div></div>'
                for label, value in metrics
            )
            flow_cards.append(
                f'<div style="border:1px solid #e4e7ec;border-top:4px solid {status_color};border-radius:15px;'
                f'padding:16px;background:white;box-shadow:0 3px 12px rgba(16,24,40,.04);">'
                f'<div style="display:flex;justify-content:space-between;gap:10px;align-items:start;">'
                f'<div><div style="font-size:17px;font-weight:900;color:#101828;">{html.escape(str(row["품목"]))}</div>'
                f'<div style="font-size:11px;color:#667085;margin-top:3px;">관련 구매품 · {html.escape(str(row["관련 구매품"]))}</div></div>'
                f'<span title="{html.escape(interpretation)}" style="padding:5px 9px;border-radius:999px;background:{status_color}12;color:{status_color};'
                f'font-size:11px;font-weight:850;white-space:nowrap;">{status}</span></div>'
                f'<div style="font-size:23px;font-weight:900;margin:14px 0 11px;color:#101828;">'
                f'{row["최신값"]:,.2f} <span style="font-size:13px;color:#667085;">{html.escape(display_unit)}</span></div>'
                f'<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:6px;">{metric_html}</div>'
                f'<div style="font-size:12px;color:#475467;line-height:1.55;margin-top:11px;">{interpretation}</div></div>'
            )
        st.markdown(
            '<div style="display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px;">'
            + "".join(flow_cards) + '</div>', unsafe_allow_html=True,
        )

    with beta_tabs[2]:
        st.markdown("#### 최근 12개월 가격 추이")
        st.caption("최근 12개월 동안 품목별 가격이 어떻게 움직였는지 보여줍니다.")
        if history_all is None or history_all.empty:
            st.info("스파크라인에 사용할 데이터가 없습니다.")
        else:
            cols = st.columns(2)
            for index, (_, row) in enumerate(raw.iterrows()):
                history = history_all[history_all["item"] == row["품목"]].sort_values("date").tail(12)
                if history.empty:
                    continue
                fig = go.Figure(go.Scatter(
                    x=history["date"], y=history["value"], mode="lines+markers",
                    line=dict(color=tone(row["전월 증감률"]), width=3), marker=dict(size=5),
                    hovertemplate="%{x|%Y년 %m월}<br>%{y:,.2f}<extra></extra>",
                ))
                fig.update_layout(
                    height=220, margin=dict(l=10, r=10, t=50, b=25), showlegend=False,
                    title=dict(text=f'<b>{row["품목"]}</b> · 전월 {pct(row["전월 증감률"])}', x=.03),
                    xaxis=dict(showgrid=False, tickformat="%y.%m", nticks=6),
                    yaxis=dict(gridcolor="#eef1f5"), plot_bgcolor="white", paper_bgcolor="white",
                )
                cols[index % 2].plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

    with beta_tabs[3]:
        st.markdown("#### 현재 가격 위치")
        st.caption("현재 가격이 최근 12개월의 최저·평균·최고 중 어느 위치인지 보여줍니다.")
        cards = []
        if history_all is not None and not history_all.empty:
            for _, row in raw.iterrows():
                history = history_all[history_all["item"] == row["품목"]].sort_values("date").tail(12)
                if history.empty:
                    continue
                low, high = float(history["value"].min()), float(history["value"].max())
                average, latest = float(history["value"].mean()), float(history.iloc[-1]["value"])
                position = 50 if high == low else (latest - low) / (high - low) * 100
                average_position = 50 if high == low else (average - low) / (high - low) * 100
                accent = tone(row["전월 증감률"])
                cards.append(
                    f'<div style="border:1px solid #e4e7ec;border-radius:14px;padding:16px;background:white;">'
                    f'<div style="display:flex;justify-content:space-between;"><b>{html.escape(str(row["품목"]))}</b><b style="font-size:19px;color:{accent};">{latest:,.2f}</b></div>'
                    f'<div style="position:relative;height:12px;border-radius:8px;margin:24px 6px 10px;background:linear-gradient(90deg,#dcecff,#f4f5f7,#ffe0e0);">'
                    f'<span style="position:absolute;left:{average_position:.1f}%;top:-14px;transform:translateX(-50%);font-size:9px;color:#667085;">평균</span>'
                    f'<span style="position:absolute;left:{position:.1f}%;top:50%;transform:translate(-50%,-50%);width:18px;height:18px;border:4px solid white;border-radius:50%;background:{accent};box-shadow:0 1px 5px #999;"></span></div>'
                    f'<div style="display:flex;justify-content:space-between;font-size:11px;color:#667085;"><span>최저 {low:,.2f}</span><span>평균 {average:,.2f}</span><span>최고 {high:,.2f}</span></div>'
                    f'<div style="margin-top:9px;color:{accent};font-weight:800;">전월 {pct(row["전월 증감률"])}</div></div>'
                )
        st.markdown('<div style="display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px;">' + "".join(cards) + '</div>', unsafe_allow_html=True)


def render_summary_table(df, show_related=True):
    snapshot = get_snapshot(df)
    metadata = get_item_meta(df)

    rows = []
    for _, row in snapshot.iterrows():
        meta = metadata.get(row["item"], {})
        currency = meta.get("currency") or "-"
        unit = meta.get("unit") or "-"
        value_unit = combined_unit(meta.get("currency"), meta.get("unit"))
        summary_row = {
            "품목": row["item"],
            "통화": currency,
            "단위": unit,
            "최신 기준월": f"{row['latest_date'].year}년 {row['latest_date'].month}월",
            "최신값": format_value(row["latest_value"], value_unit),
            "전월 가격": format_value(row["prev_value"], value_unit),
            "전월 차이": format_signed_value(row["mom_difference"], value_unit),
            "전월 증감률": format_delta(row["mom"]),
            "6개월 전 가격": format_value(
                row["six_month_value"], value_unit
            ),
            "6개월 차이": format_signed_value(
                row["six_month_difference"], value_unit
            ),
            "6개월 증감률": format_delta(row["six_month_change"]),
            f"{int(row['previous_year'])}년 평균가격": format_value(
                row["previous_year_average"], value_unit
            ),
            "전년 차이": format_signed_value(
                row["previous_year_difference"], value_unit
            ),
            "전년 증감률": format_delta(row["previous_year_change"]),
            "출처": meta.get("source") or "-",
        }
        if show_related:
            summary_row = {
                "품목": summary_row.pop("품목"),
                "사용 품목": meta.get("related") or "-",
                **summary_row,
            }
        rows.append(summary_row)

    summary_df = pd.DataFrame(rows)
    change_columns = [
        "전월 차이", "전월 증감률",
        "6개월 차이", "6개월 증감률",
        "전년 차이", "전년 증감률",
    ]
    summary_html = map_dataframe(summary_df, lambda value: html.escape(str(value)))
    for column in change_columns:
        if column in summary_df.columns:
            summary_html[column] = summary_df[column].map(colorize_change)

    table_html = summary_html.to_html(index=False, escape=False, classes="comparison-table")
    comparison_css = (
        "<style>"
        ".comparison-scroll{overflow-x:auto;margin-bottom:.5rem}"
        ".comparison-table{width:100%;border-collapse:collapse;font-size:.92rem}"
        ".comparison-table th{background:#f3f5f8;color:#263238;font-weight:700;"
        "padding:10px 9px;border:1px solid #dfe3e8;white-space:nowrap}"
        ".comparison-table td{padding:9px;border:1px solid #e5e7eb;white-space:nowrap}"
        ".comparison-table tbody tr:hover{background:#fafbfc}"
        ".change-up{display:inline-block;color:#c62828;font-weight:800}"
        ".change-down{display:inline-block;color:#1565c0;font-weight:800}"
        ".change-flat{display:inline-block;color:#616161;font-weight:700}"
        "</style>"
    )
    st.markdown(
        comparison_css + '<div class="comparison-scroll">' + table_html + "</div>",
        unsafe_allow_html=True,
    )


def render_monthly_table(df, show_related=True, comparison_df=None, key="monthly"):
    comparison_source = comparison_df if comparison_df is not None else df
    metadata = get_item_meta(comparison_source)
    snapshot = get_snapshot(comparison_source).set_index("item")
    available_months = df["date"].drop_duplicates().sort_values().tolist()
    if not available_months:
        st.info("월별 표에 표시할 데이터가 없습니다.")
        return

    view_mode = st.radio(
        "표 보기",
        options=["간단히", "자세히"],
        horizontal=True,
        key=f"{key}_table_view_mode",
        help="간단히는 핵심 증감률만, 자세히는 차이와 전년도 평균까지 표시합니다.",
    )
    default_start = available_months[max(0, len(available_months) - 6)]

    table_start, table_end = st.select_slider(
        "월별 데이터 조회 기간",
        options=available_months,
        value=(default_start, available_months[-1]),
        format_func=lambda value: f"{value.year}년 {value.month}월",
        key=f"{key}_table_period_v2",
    )
    table_df = df[df["date"].between(table_start, table_end)].copy()
    month_dates = table_df["date"].drop_duplicates().sort_values().tolist()
    month_order = [f"{value.year}년 {value.month}월" for value in month_dates]
    month_headers = {
        label: (f"{value.year}년", f"{value.month}월")
        for label, value in zip(month_order, month_dates)
    }

    pivot = (
        table_df.pivot_table(index="item", columns="ym", values="value", aggfunc="last")
        .reindex(columns=month_order)
        .reset_index()
        .rename(columns={"item": "품목"})
    )

    if show_related:
        pivot.insert(
            1,
            "관련 구매품",
            pivot["품목"].map(lambda x: metadata.get(x, {}).get("related") or "-"),
        )
    currency_position = 2 if show_related else 1
    pivot.insert(
        currency_position,
        "통화",
        pivot["품목"].map(lambda x: metadata.get(x, {}).get("currency") or "-"),
    )
    pivot.insert(
        currency_position + 1,
        "단위",
        pivot["품목"].map(lambda x: metadata.get(x, {}).get("unit") or "-"),
    )
    pivot["전월 차이"] = pivot["품목"].map(
        lambda item: snapshot.loc[item, "mom_difference"] if item in snapshot.index else pd.NA
    )
    pivot["전월 증감률"] = pivot["품목"].map(
        lambda item: snapshot.loc[item, "mom"] if item in snapshot.index else pd.NA
    )
    pivot["전년 평균가격"] = pivot["품목"].map(
        lambda item: snapshot.loc[item, "previous_year_average"] if item in snapshot.index else pd.NA
    )
    pivot["전년 차이"] = pivot["품목"].map(
        lambda item: snapshot.loc[item, "previous_year_difference"] if item in snapshot.index else pd.NA
    )
    pivot["전년 증감률"] = pivot["품목"].map(
        lambda item: snapshot.loc[item, "previous_year_change"] if item in snapshot.index else pd.NA
    )

    info_columns = ["품목"]
    if show_related:
        info_columns.append("관련 구매품")
    info_columns.extend(["통화", "단위"])
    previous_years = snapshot["previous_year"].dropna()
    previous_year_label = (
        f"{int(previous_years.max())}년 평균가격"
        if not previous_years.empty
        else "전년도 평균가격"
    )

    if view_mode == "간단히":
        comparison_columns = ["전월 증감률", "전년 증감률"]
        comparison_headers = [
            ("전월 대비", "증감률"),
            ("전년 대비", "증감률"),
        ]
    else:
        comparison_columns = [
            "전월 차이", "전월 증감률", "전년 평균가격", "전년 차이", "전년 증감률"
        ]
        comparison_headers = [
            ("전월 대비", "차이"),
            ("전월 대비", "증감률"),
            ("전년 대비", previous_year_label),
            ("전년 대비", "차이"),
            ("전년 대비", "증감률"),
        ]

    display_columns = info_columns + month_order + comparison_columns
    monthly_view = pivot[display_columns].copy()
    monthly_view.columns = pd.MultiIndex.from_tuples(
        [("기본 정보", column) for column in info_columns]
        + [month_headers[column] for column in month_order]
        + comparison_headers
    )

    safe_view = map_dataframe(
        monthly_view,
        lambda value: "-" if pd.isna(value) else html.escape(str(value))
    )
    safe_view[("기본 정보", "품목")] = monthly_view[("기본 정보", "품목")].map(
        lambda item: (
            f'<strong>{html.escape(str(item))}</strong>'
            + (
                f'<div class="item-note">{html.escape(str(metadata.get(item, {}).get("note") or ""))}</div>'
                if metadata.get(item, {}).get("note") else ""
            )
        )
    )
    for month in month_order:
        month_header = month_headers[month]
        safe_view[month_header] = monthly_view[month_header].map(
            lambda value: "-" if pd.isna(value) else f"{value:,.2f}"
        )
    safe_view[("전월 대비", "증감률")] = monthly_view[("전월 대비", "증감률")].map(
        lambda value: colorize_change(format_delta(value))
    )
    safe_view[("전년 대비", "증감률")] = monthly_view[("전년 대비", "증감률")].map(
        lambda value: colorize_change(format_delta(value))
    )
    if view_mode == "자세히":
        safe_view[("전월 대비", "차이")] = monthly_view[("전월 대비", "차이")].map(
            lambda value: colorize_change(format_signed_value(value))
        )
        safe_view[("전년 대비", previous_year_label)] = monthly_view[
            ("전년 대비", previous_year_label)
        ].map(lambda value: "-" if pd.isna(value) else f"{value:,.2f}")
        safe_view[("전년 대비", "차이")] = monthly_view[("전년 대비", "차이")].map(
            lambda value: colorize_change(format_signed_value(value))
        )

    year_groups = []
    for month_date in month_dates:
        year_label = f"{month_date.year}년"
        if year_groups and year_groups[-1][0] == year_label:
            year_groups[-1] = (year_label, year_groups[-1][1] + 1)
        else:
            year_groups.append((year_label, 1))

    top_headers = "".join(
        f'<th rowspan="2">{html.escape(column)}</th>' for column in info_columns
    )
    top_headers += "".join(
        f'<th colspan="{count}">{year_label}</th>'
        for year_label, count in year_groups
    )
    if view_mode == "간단히":
        top_headers += '<th colspan="1">전월 대비</th><th colspan="1">전년 대비</th>'
    else:
        top_headers += '<th colspan="2">전월 대비</th><th colspan="3">전년 대비</th>'

    sub_headers = "".join(
        (
            f'<th class="latest-month">{month_date.month}월</th>'
            if month_date == month_dates[-1]
            else f"<th>{month_date.month}월</th>"
        )
        for month_date in month_dates
    )
    if view_mode == "간단히":
        sub_headers += "<th>증감률</th><th>증감률</th>"
    else:
        sub_headers += (
            "<th>차이</th><th>증감률</th>"
            f"<th>{html.escape(previous_year_label)}</th><th>차이</th><th>증감률</th>"
        )
    safe_rows = list(safe_view.itertuples(index=False, name=None))
    fixed_columns = ["품목"] + (["관련 구매품"] if show_related else [])
    fixed_count = len(fixed_columns)
    scroll_info_columns = ["통화", "단위"]
    fixed_headers = "".join(
        f"<th>{html.escape(column)}</th>" for column in fixed_columns
    )
    fixed_rows = "".join(
        "<tr>" + "".join(f"<td>{value}</td>" for value in row[:fixed_count]) + "</tr>"
        for row in safe_rows
    )
    latest_month_cell_index = len(scroll_info_columns) + len(month_order) - 1
    data_rows = "".join(
        "<tr>"
        + "".join(
            (
                f'<td class="latest-month">{value}</td>'
                if index == latest_month_cell_index
                else f"<td>{value}</td>"
            )
            for index, value in enumerate(row[fixed_count:])
        )
        + "</tr>"
        for row in safe_rows
    )
    right_top_headers = "".join(
        f'<th rowspan="2">{html.escape(column)}</th>'
        for column in scroll_info_columns
    )
    right_top_headers += "".join(
        f'<th colspan="{count}">{year_label}</th>'
        for year_label, count in year_groups
    )
    if view_mode == "간단히":
        right_top_headers += '<th colspan="1">전월 대비</th><th colspan="1">전년 대비</th>'
    else:
        right_top_headers += '<th colspan="2">전월 대비</th><th colspan="3">전년 대비</th>'
    fixed_width = 370 if show_related else 150
    table_html = (
        '<div class="monthly-table-shell">'
        f'<div class="monthly-fixed-pane" style="flex-basis:{fixed_width}px;min-width:{fixed_width}px">'
        '<table class="monthly-comparison-table fixed-table">'
        f"<thead><tr>{fixed_headers}</tr></thead><tbody>{fixed_rows}</tbody></table></div>"
        '<div class="monthly-scroll-pane"><table class="monthly-comparison-table data-table">'
        f"<thead><tr>{right_top_headers}</tr><tr>{sub_headers}</tr></thead>"
        f"<tbody>{data_rows}</tbody></table></div></div>"
    )
    monthly_css = (
        "<style>"
        ".monthly-table-shell{display:flex;width:100%;border:1px solid #d9dee7;"
        "border-radius:10px;overflow:hidden;background:#fff}"
        ".monthly-fixed-pane{flex:0 0 auto;position:relative;z-index:3;"
        "box-shadow:5px 0 8px -5px #697786;background:#fff}"
        ".monthly-scroll-pane{flex:1 1 auto;min-width:0;overflow-x:auto}"
        ".monthly-comparison-table{width:max-content;min-width:100%;border-collapse:collapse;"
        "font-size:1rem;color:#17233b}"
        ".monthly-comparison-table th{background:#e9edf3;color:#17233b;font-weight:800;"
        "padding:10px 9px;border:1px solid #cfd6df;white-space:nowrap;text-align:center}"
        ".monthly-comparison-table thead tr:first-child th{background:#dce3ec;"
        "font-size:1.02rem;border-bottom:2px solid #aeb8c6}"
        ".fixed-table{width:100%;table-layout:fixed}"
        ".fixed-table thead th{height:82px;vertical-align:middle;background:#dce3ec!important}"
        ".data-table thead tr th{height:41px;box-sizing:border-box}"
        ".monthly-comparison-table tbody tr{height:52px}"
        ".monthly-comparison-table tbody td{height:52px;box-sizing:border-box}"
        ".fixed-table th:nth-child(1),.fixed-table td:nth-child(1){width:150px}"
        ".fixed-table th:nth-child(2),.fixed-table td:nth-child(2){width:220px}"
        ".monthly-comparison-table td{padding:10px 9px;border:1px solid #e0e4ea;"
        "white-space:nowrap;text-align:right;font-weight:520}"
        ".monthly-comparison-table td:first-child{text-align:left;font-weight:750}"
        ".monthly-comparison-table .item-note{font-size:.76rem;color:#7a8391;"
        "font-weight:500;margin-top:3px;text-align:left}"
        ".monthly-comparison-table th.latest-month{background:#fff0c7!important;"
        "color:#7a4a00!important;border-color:#e5bd62!important}"
        ".monthly-comparison-table td.latest-month{background:#fffaf0!important;"
        "font-weight:800;color:#17233b;border-left:2px solid #e5bd62}"
        ".monthly-comparison-table tbody tr:nth-child(even){background:#f8fafc}"
        ".monthly-comparison-table tbody tr:hover{background:#fff8e6}"
        ".monthly-comparison-table .change-up{color:#c62828;font-weight:850}"
        ".monthly-comparison-table .change-down{color:#1565c0;font-weight:850}"
        ".monthly-comparison-table .change-flat{color:#616161;font-weight:750}"
        "</style>"
    )
    st.markdown(
        monthly_css + table_html,
        unsafe_allow_html=True,
    )
    st.caption("노란색 열은 선택 기간의 최신월입니다. 빨간색은 상승, 파란색은 하락을 의미합니다.")


def render_chart(df, key):
    available_months = df["date"].drop_duplicates().sort_values().tolist()
    if not available_months:
        st.info("차트에 표시할 월별 데이터가 없습니다.")
        return

    chart_start, chart_end = st.select_slider(
        "차트 조회 기간",
        options=available_months,
        value=(available_months[0], available_months[-1]),
        format_func=lambda value: f"{value.year}년 {value.month}월",
        key=f"{key}_chart_period",
    )
    chart_source = df[df["date"].between(chart_start, chart_end)].copy()

    items = sorted(chart_source["item"].unique().tolist())
    state_key = f"{key}_chart_selected_items"
    if state_key not in st.session_state:
        st.session_state[state_key] = items[: min(4, len(items))]
    st.session_state[state_key] = [item for item in st.session_state[state_key] if item in items]
    st.markdown("**차트 품목 선택**")
    st.caption("품목 버튼을 누르면 차트에 추가되거나 제외됩니다.")
    left, right, _ = st.columns([1, 1, 6])
    if left.button("전체 선택", key=f"{key}_select_all", use_container_width=True):
        st.session_state[state_key] = items.copy()
        st.rerun()
    if right.button("선택 해제", key=f"{key}_clear_all", use_container_width=True):
        st.session_state[state_key] = []
        st.rerun()
    button_cols = st.columns(min(6, max(1, len(items))))
    for index, item in enumerate(items):
        is_selected = item in st.session_state[state_key]
        if button_cols[index % len(button_cols)].button(
            f'{"✓" if is_selected else "+"} {item}',
            key=f"{key}_item_button_{index}",
            type="primary" if is_selected else "secondary",
            use_container_width=True,
        ):
            if is_selected:
                st.session_state[state_key].remove(item)
            else:
                st.session_state[state_key].append(item)
            st.rerun()
    selected = st.session_state[state_key]

    if not selected:
        st.info("차트에 표시할 품목을 1개 이상 선택해 주세요.")
        return

    normalize = st.toggle(
        "변동률 비교용 지수화 (첫 달=100)",
        value=False,
        key=f"{key}_normalize",
        help="단위가 서로 다른 원자재나 환율의 움직임을 한 차트에서 비교할 때 사용합니다.",
    )
    show_moving_average = st.toggle(
        "3개월 추세선 표시",
        value=False,
        key=f"{key}_moving_average",
        help="월별 변동을 완화해 최근 방향을 확인하는 3개월 이동평균입니다.",
    )

    chart_df = chart_source[chart_source["item"].isin(selected)].copy()
    if "currency" not in chart_df.columns:
        chart_df["currency"] = ""
    chart_df["display_unit"] = chart_df.apply(
        lambda row: combined_unit(row.get("currency", ""), row.get("unit", "")),
        axis=1,
    )

    if normalize:
        chart_df["chart_value"] = chart_df.groupby("item")["value"].transform(
            lambda x: x / x.iloc[0] * 100 if x.iloc[0] != 0 else x
        )
        y_title = "지수 (첫 달=100)"
    else:
        chart_df["chart_value"] = chart_df["value"]
        y_title = "값"

    fig = go.Figure()
    line_colors = [
        "#1565C0", "#EF5350", "#2E7D32", "#8E24AA",
        "#F57C00", "#00838F", "#6D4C41", "#C2185B",
    ]
    label_endpoints = []

    for item_index, item in enumerate(selected):
        item_df = chart_df[chart_df["item"] == item].sort_values("date")
        line_color = line_colors[item_index % len(line_colors)]
        fig.add_trace(
            go.Scatter(
                x=item_df["date"],
                y=item_df["chart_value"],
                mode="lines+markers",
                name=item,
                line={"width": 3, "color": line_color},
                marker={"size": 7, "color": line_color},
                customdata=item_df[["value", "display_unit"]],
                hovertemplate=(
                    "<b>%{fullData.name}</b><br>"
                    "기준월: %{x|%Y년 %m월}<br>"
                    "값: %{customdata[0]:,.2f} %{customdata[1]}"
                    "<extra></extra>"
                ),
            )
        )
        if show_moving_average:
            moving_average = item_df["chart_value"].rolling(3, min_periods=2).mean()
            fig.add_trace(
                go.Scatter(
                    x=item_df["date"],
                    y=moving_average,
                    mode="lines",
                    name=f"{item} 3개월 추세",
                    line={"width": 2, "color": line_color, "dash": "dot"},
                    opacity=0.8,
                    showlegend=False,
                    hovertemplate=(
                        f"<b>{html.escape(str(item))} 3개월 추세</b><br>"
                        "기준월: %{x|%Y년 %m월}<br>값: %{y:,.2f}<extra></extra>"
                    ),
                )
            )
        if not item_df.empty:
            last_row = item_df.iloc[-1]
            label_endpoints.append({
                "item": item,
                "date": last_row["date"],
                "value": float(last_row["chart_value"]),
                "color": line_color,
            })

    # 끝점 값이 가까운 지표들은 라벨을 위아래로 자동 분산합니다.
    if label_endpoints:
        endpoint_values = [point["value"] for point in label_endpoints]
        value_range = max(endpoint_values) - min(endpoint_values)
        collision_threshold = max(value_range * 0.06, abs(max(endpoint_values)) * 0.008, 1e-9)
        sorted_points = sorted(label_endpoints, key=lambda point: point["value"])
        label_groups = []
        current_group = [sorted_points[0]]

        for point in sorted_points[1:]:
            if point["value"] - current_group[-1]["value"] <= collision_threshold:
                current_group.append(point)
            else:
                label_groups.append(current_group)
                current_group = [point]
        label_groups.append(current_group)

        for group in label_groups:
            group_size = len(group)
            for group_index, point in enumerate(group):
                vertical_offset = int((group_index - (group_size - 1) / 2) * 34)
                fig.add_annotation(
                    x=point["date"],
                    y=point["value"],
                    text=f"<b>{html.escape(str(point['item']))}</b>",
                    showarrow=True,
                    arrowhead=0,
                    arrowwidth=2,
                    arrowcolor=point["color"],
                    ax=42,
                    ay=vertical_offset,
                    xanchor="left",
                    font={"color": point["color"], "size": 13},
                    bgcolor="rgba(255,255,255,0.94)",
                    bordercolor=point["color"],
                    borderwidth=1,
                    borderpad=4,
                )

    fig.update_layout(
        height=520,
        margin={"l": 20, "r": 230, "t": 20, "b": 20},
        hovermode="x unified",
        legend={
            "orientation": "h",
            "yanchor": "bottom",
            "y": 1.02,
            "xanchor": "left",
            "x": 0,
        },
        xaxis={"title": "", "tickformat": "%Y년 %m월"},
        yaxis={"title": y_title},
    )

    st.plotly_chart(fig, use_container_width=True)


def render_colored_kpis(df, show_related=True):
    snapshot = get_snapshot(df)
    metadata = get_item_meta(df)
    cards = []
    for _, row in snapshot.iterrows():
        meta = metadata.get(row["item"], {})
        mom = row["mom"]
        if pd.isna(mom) or abs(mom) < 0.01:
            color = "flat"
            delta_text = "전월 데이터 없음" if pd.isna(mom) else "전월 대비 ― 0.0%"
        elif mom > 0:
            color = "up"
            delta_text = f"전월 대비 ▲ {abs(mom):.1f}%"
        else:
            color = "down"
            delta_text = f"전월 대비 ▼ {abs(mom):.1f}%"
        value_unit = combined_unit(meta.get("currency"), meta.get("unit"))
        related = meta.get("related") or "-"
        note = str(meta.get("note") or "").strip()
        note_html = (
            f'<div class="briefing-card-note">{html.escape(note)}</div>'
            if note and show_related else ""
        )
        sub_text = (
            f"관련 구매품: {related}"
            if show_related
            else f"{row['latest_date'].year}년 {row['latest_date'].month}월 기준"
        )
        month_text = (
            f"{row['latest_date'].year}년 {row['latest_date'].month}월 · "
            if show_related
            else ""
        )
        cards.append(
            f'<div class="briefing-card {color}">'
            f'<div class="briefing-card-name">{html.escape(str(row["item"]))}</div>'
            f'{note_html}'
            f'<div class="briefing-card-related">{html.escape(sub_text)}</div>'
            f'<span class="briefing-card-value">{html.escape(format_value(row["latest_value"], value_unit))}</span>'
            f'<span class="briefing-card-change {color}">{html.escape(delta_text)}</span>'
            f'<div class="briefing-card-note">{html.escape(month_text)}최근 월 기준</div>'
            '</div>'
        )
    st.markdown(
        f'<div class="briefing-grid">{"".join(cards)}</div>',
        unsafe_allow_html=True,
    )


def render_section(df, key, months_to_show):
    if df.empty:
        st.warning("표시할 데이터가 없습니다.")
        return

    filtered = filter_period(df, months_to_show)
    show_related = key != "환율"

    latest = filtered["date"].max()
    earliest = filtered["date"].min()
    st.caption(
        f"표시 기간: {earliest.year}년 {earliest.month}월 ~ "
        f"{latest.year}년 {latest.month}월 "
        f"/ {filtered['item'].nunique()}개 품목"
    )

    # 원자재·환율 탭의 기존 카드형 요약 지표를 유지합니다.
    section_analysis = build_procurement_metrics({key: df})
    if not section_analysis.empty:
        st.markdown("### 요약 지표")
        render_item_briefing_cards(
            section_analysis,
            show_related=(key == "원자재"),
            show_trend=False,
        )

    st.markdown("### 원자재 가격 추이" if key == "원자재" else "### 환율 추이")
    # 차트 기간은 업로드된 전체 월 범위에서 별도로 선택합니다.
    render_chart(df, key)

    st.divider()
    st.markdown("### 월별 상세 데이터")
    render_monthly_table(
        df,
        show_related=show_related,
        comparison_df=df,
        key=key,
    )


# =========================================================
# 엑셀 템플릿 생성
# =========================================================
def make_template_bytes():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    current_year = NOW.year
    previous_year = current_year - 1
    raw_items = [
        ("", "알루미늄", "알루미늄리드지", "USD", "ton"),
        ("", "두바이유", "PP수지, 블리스터", "USD", "bbl"),
        ("", "나프타", "PP수지, HEMA, GMMA 등", "USD", "MT"),
        ("", "에틸렌", "BE", "USD", "MT"),
        ("종가 기준", "옥수수", "알코올", "USD", "ton"),
        ("월 말 업데이트", "펄프", "세일즈팩", "USD", "ton"),
        ("월 말 업데이트", "폐골판지", "박스류", "KRW", "kg"),
    ]
    fx_items = [
        ("", "USD", "", "KRW", "1 USD"),
        ("", "EUR", "", "KRW", "1 EUR"),
        ("", "JPY", "", "KRW", "100 JPY"),
    ]
    source_rows = [
        ("알루미늄", "알루미늄리드지", "USD", "ton", "산업통상부", "https://www.motir.go.kr/kor/contents/103", "일간"),
        ("두바이유", "PP수지, 블리스터", "USD", "bbl", "산업통상부", "https://www.motir.go.kr/kor/contents/103", "일간"),
        ("나프타", "PP수지, HEMA, GMMA 등", "USD", "MT", "산업통상부", "https://www.motir.go.kr/kor/contents/103", "주간"),
        ("에틸렌", "BE", "USD", "MT", "산업통상부", "https://www.motir.go.kr/kor/contents/103", "주간"),
        ("옥수수", "알코올", "USD", "ton", "농넷(한국농수산식품유통공사)", "https://www.nongnet.or.kr/front/M000000270/globalGrain/entire.do", "일간"),
        ("펄프", "세일즈팩", "USD", "ton", "산업통상부", "https://www.motir.go.kr/kor/contents/103", "월간"),
        ("폐골판지", "박스류", "KRW", "kg", "산업통상부", "https://www.motir.go.kr/kor/contents/103", "월간"),
    ]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"{current_year}년"
    source_sheet = workbook.create_sheet("출처")

    thin_gray = Side(style="thin", color="B7B7B7")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_fill = PatternFill("solid", fgColor="E7E6E6")
    year_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="DDEBF7")

    worksheet.merge_cells(start_row=2, start_column=7, end_row=2, end_column=18)
    worksheet.merge_cells(start_row=2, start_column=19, end_row=2, end_column=30)
    worksheet.cell(2, 7, f"{previous_year}년")
    worksheet.cell(2, 19, f"{current_year}년")
    for cell in [worksheet.cell(2, 7), worksheet.cell(2, 19)]:
        cell.fill = year_fill
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["비고", "no.", "구분", "관련 구매품", "통화", "단위"]
    for column, header in enumerate(headers, start=1):
        worksheet.cell(3, column, header)
    month_column = 7
    for year in [previous_year, current_year]:
        for month in range(1, 13):
            cell = worksheet.cell(3, month_column, f"{year}년 {month}월")
            cell.number_format = "yyyy년 m월"
            month_column += 1

    for column in range(1, 31):
        cell = worksheet.cell(3, column)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    all_items = raw_items + fx_items
    for row_number, (note, item, related, currency, unit) in enumerate(all_items, start=4):
        values = [note, row_number - 3, item, related, currency, unit]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row_number, column, value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if column not in [3, 4] else "left",
                vertical="center",
            )
        for column in range(7, 31):
            worksheet.cell(row_number, column).border = border
            worksheet.cell(row_number, column).number_format = "#,##0.00"

    worksheet.freeze_panes = "G4"
    worksheet.auto_filter.ref = f"A3:AD{3 + len(all_items)}"
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 6
    worksheet.column_dimensions["C"].width = 28
    worksheet.column_dimensions["D"].width = 28
    worksheet.column_dimensions["E"].width = 10
    worksheet.column_dimensions["F"].width = 11
    for column in range(7, 31):
        worksheet.column_dimensions[get_column_letter(column)].width = 13
    worksheet.row_dimensions[2].height = 22
    worksheet.row_dimensions[3].height = 30

    source_sheet.cell(2, 2, "▶ 원자재")
    source_sheet.cell(2, 2).fill = section_fill
    source_headers = ["no.", "원료항목", "관련 구매품", "통화", "단위", "출처", "URL", "주기"]
    for column, header in enumerate(source_headers, start=2):
        source_sheet.cell(3, column, header)
    for row_number, row in enumerate(source_rows, start=4):
        values = [row_number - 3] + list(row)
        for column, value in enumerate(values, start=2):
            source_sheet.cell(row_number, column, value)

    fx_start = 4 + len(source_rows) + 1
    source_sheet.cell(fx_start, 2, "▶ 환율")
    source_sheet.cell(fx_start, 2).fill = section_fill
    fx_source_headers = ["no.", "통화", "출처", "비고", "URL"]
    for column, header in enumerate(fx_source_headers, start=2):
        source_sheet.cell(fx_start + 1, column, header)
    fx_sources = [
        ("유로 (EUR)", "서울외국환중개", "", "http://www.smbs.biz/ExRate/StdExRate.jsp"),
        ("미국 달러 (USD)", "서울외국환중개", "", "http://www.smbs.biz/ExRate/StdExRate.jsp"),
        ("일본 엔 (JPY)", "서울외국환중개", "100 JPY 기준", "http://www.smbs.biz/ExRate/StdExRate.jsp"),
    ]
    for row_number, row in enumerate(fx_sources, start=fx_start + 2):
        values = [row_number - (fx_start + 1)] + list(row)
        for column, value in enumerate(values, start=2):
            source_sheet.cell(row_number, column, value)

    for row in source_sheet.iter_rows(min_row=3, max_row=source_sheet.max_row, min_col=2, max_col=9):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row in [3, fx_start + 1]:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
    source_sheet.freeze_panes = "B4"
    source_sheet.column_dimensions["A"].width = 2
    source_sheet.column_dimensions["B"].width = 7
    source_sheet.column_dimensions["C"].width = 28
    source_sheet.column_dimensions["D"].width = 28
    source_sheet.column_dimensions["E"].width = 11
    source_sheet.column_dimensions["F"].width = 11
    source_sheet.column_dimensions["G"].width = 28
    source_sheet.column_dimensions["H"].width = 55
    source_sheet.column_dimensions["I"].width = 12

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


# =========================================================
# 앱 본문
# =========================================================
st.title("원가 영향 인자 대시보드")
st.caption("엑셀 파일을 업로드하면 원자재 및 환율 데이터를 자동으로 분석합니다.")

with st.sidebar:
    st.header("데이터 설정")

    template_bytes = make_template_bytes()
    st.download_button(
        "엑셀 입력 양식 다운로드",
        data=template_bytes,
        file_name="원가_영향인자_입력양식.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help="처음 사용하는 사용자가 입력용 엑셀 양식을 내려받을 수 있습니다.",
    )

    uploaded_file = st.file_uploader(
        "원가 영향인자 엑셀 업로드",
        type=["xlsx", "xlsm"],
        help="권장 시트명: 원자재, 환율 / 필수 열: 품목, 날짜, 값",
    )

    storage_ready = get_storage_settings() is not None
    if uploaded_file is not None:
        save_clicked = st.button(
            "현재 파일 저장",
            type="primary",
            use_container_width=True,
            disabled=not storage_ready,
            help="이 파일을 최신 데이터로 저장합니다.",
        )
    else:
        save_clicked = False

    if not storage_ready:
        st.caption("Streamlit 설정에 GitHub 토큰을 등록하면 저장 버튼을 사용할 수 있습니다.")

    months_to_show = st.selectbox(
        "표시 기간",
        options=[6, 12, 18, 24, 36],
        index=1,
        format_func=lambda x: f"최근 {x}개월",
    )

saved_file_error = None
saved_file_bytes = None
saved_file_timestamp = None
try:
    # 공개 저장소의 최근 데이터는 토큰 없이도 조회합니다.
    # GitHub 토큰은 새 파일을 저장할 때만 필요합니다.
    saved_file_bytes = load_saved_file()
    if saved_file_bytes is not None:
        try:
            saved_file_timestamp = load_saved_timestamp()
        except Exception:
            saved_file_timestamp = None
except Exception as exc:
    saved_file_error = str(exc)

if saved_file_error:
    st.warning(f"최근 저장 파일을 불러오지 못했습니다: {saved_file_error}")

if uploaded_file is None and saved_file_bytes is None:
    st.info("왼쪽에서 엑셀 파일을 업로드해 주세요. 저장한 파일이 있으면 다음 접속부터 자동으로 표시됩니다.")
    st.stop()

try:
    file_bytes = uploaded_file.getvalue() if uploaded_file is not None else saved_file_bytes
    datasets, load_messages, sheet_names = load_excel(file_bytes)
except Exception as exc:
    st.error(f"엑셀 파일을 읽는 중 오류가 발생했습니다: {exc}")
    st.stop()

if save_clicked:
    try:
        save_latest_file(file_bytes)
        saved_file_bytes = file_bytes
        saved_file_timestamp = datetime.now(pytz.timezone("Asia/Seoul"))
        st.toast("최신 파일로 저장했습니다.", icon="✅")
        st.success(
            "저장이 완료되었습니다. "
            f"최근 저장 일시: {saved_file_timestamp.strftime('%Y년 %m월 %d일 %H:%M')}"
        )
    except Exception as exc:
        st.error(f"파일을 저장하지 못했습니다: {exc}")

if saved_file_bytes is not None:
    saved_download_name = (
        "원가영향인자_최근저장본_"
        + (
            saved_file_timestamp.strftime("%Y%m%d_%H%M")
            if saved_file_timestamp is not None
            else "latest"
        )
        + ".xlsx"
    )
    with st.sidebar:
        st.download_button(
            "최근 저장 파일 다운로드",
            data=saved_file_bytes,
            file_name=saved_download_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="GitHub에 마지막으로 저장된 엑셀 파일을 내려받습니다.",
        )

if uploaded_file is not None:
    st.success(
        f"'{uploaded_file.name}' 파일을 불러왔습니다. "
        f"확인된 시트: {', '.join(sheet_names)}"
    )
else:
    st.success(
        "최근 저장한 파일을 자동으로 불러왔습니다. "
        f"확인된 시트: {', '.join(sheet_names)}"
    )
    if saved_file_timestamp is not None:
        st.info(
            "최근 저장 일시: "
            f"{saved_file_timestamp.strftime('%Y년 %m월 %d일 %H:%M')}"
        )

for message in load_messages:
    st.warning(message)

available_tabs = []
if "원자재" in datasets and not datasets["원자재"].empty:
    available_tabs.append("원자재")
if "환율" in datasets and not datasets["환율"].empty:
    available_tabs.append("환율")

if not available_tabs:
    st.error("원자재 또는 환율 데이터를 찾지 못했습니다.")
    st.stop()

tab_names = ["월간 동향", "분석 인사이트"]
if "원자재" in available_tabs:
    tab_names.append("원자재 가격")
if "환율" in available_tabs:
    tab_names.append("환율 정보")
tabs = st.tabs(tab_names)

for tab, tab_name in zip(tabs, tab_names):
    with tab:
        if tab_name == "월간 동향":
            render_monthly_briefing(datasets)
            continue
        if tab_name == "분석 인사이트":
            render_beta_analysis(datasets)
            continue
        dataset_key = "원자재" if tab_name == "원자재 가격" else "환율"
        st.subheader(tab_name)
        render_section(
            datasets[dataset_key],
            key=dataset_key,
            months_to_show=months_to_show,
        )

st.divider()
st.caption(
    "업로드한 파일 안의 값만 사용하며, 외부 웹사이트나 API에는 자동 연결하지 않습니다."
)
